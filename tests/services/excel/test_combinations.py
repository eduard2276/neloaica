"""Combinatorial receipt generation tests.

Each section (defects, discovered_defects, client_parts, labor,
billable_parts) is tested at every cardinality from 0 up to its maximum,
while all other sections are kept at their maximum.  The _assert_structure
helper from test_excel_export is re-used so that every combination also
validates the structural labels (MANOPERA, Total manopera, PIESE, Total piese).

Section maxima used in these tests:
    defects              – 5  (template rows A14-A18)
    discovered_defects   – 5  (template rows A20-A24)
    client_parts         – 3  (template rows C14-C16)
    labor                – 3  (three entries cover N=1,2,3 edge cases)
    billable_parts       – 3  (three entries cover M=1,2,3 edge cases)
"""

import pytest
from unittest.mock import patch

from src.services.excel_export import generate_receipt_excel

# ---------------------------------------------------------------------------
# Re-use constants and helpers from the primary test module
# ---------------------------------------------------------------------------

TVA = 21.0

MOCK_DEFECTS = {
    1: {"id": 1, "defect_name": "Placute frana uzate"},
    2: {"id": 2, "defect_name": "Scurgere ulei motor"},
    3: {"id": 3, "defect_name": "Presiune scazuta anvelope"},
    4: {"id": 4, "defect_name": "Nivel lichid racire scazut"},
    5: {"id": 5, "defect_name": "Stergator parbriz uzat"},
}

MOCK_CLIENT_PARTS = {
    10: {"id": 10, "part_name": "Filtru ulei"},
    11: {"id": 11, "part_name": "Filtru aer"},
    12: {"id": 12, "part_name": "Filtru habitaclu"},
}

MOCK_LABOR = {
    100: {"id": 100, "service_name": "Schimb ulei"},
    101: {"id": 101, "service_name": "Inlocuire placute frana"},
    102: {"id": 102, "service_name": "Geometrie roti"},
}

ALL_DEFECT_IDS       = [1, 2, 3, 4, 5]
ALL_CLIENT_PART_IDS  = [10, 11, 12]
ALL_LABOR_IDS        = [100, 101, 102]
ALL_BILLABLE_PARTS   = [
    {"part_id": 10, "part_name": "Filtru ulei",      "units": 2.0,  "price_per_unit": 50.0},
    {"part_id": 11, "part_name": "Ulei motor 5W40",  "units": 1.5,  "price_per_unit": 80.0},
    {"part_id": 12, "part_name": "Filtru aer",       "units": 1.0,  "price_per_unit": 35.0},
]


def _make_receipt(
    defects=None,
    discovered_defects=None,
    client_parts=None,
    labor=None,
    billable_parts=None,
) -> dict:
    """Build a receipt dict; None means use the full list for that section."""
    if defects           is None: defects           = ALL_DEFECT_IDS
    if discovered_defects is None: discovered_defects = ALL_DEFECT_IDS
    if client_parts      is None: client_parts      = ALL_CLIENT_PART_IDS
    if labor             is None: labor             = ALL_LABOR_IDS
    if billable_parts    is None: billable_parts    = ALL_BILLABLE_PARTS

    parts_cost  = sum(p["units"] * p["price_per_unit"] for p in billable_parts)
    labor_cost  = 300.0 * len(labor)

    return {
        "client_name":       "Test Client",
        "client_address":    "Str. Test 1",
        "model":             "Test Car",
        "plate_number":      "XX00TST",
        "vin":               "TEST0000000000001",
        "kilometers":        "10000",
        "executant_name":    "Test Executant",
        "date":              "07.05.2026",
        "estimate_cost":     500.0,
        "estimated_final_date": "10.05.2026",
        "defects":           list(defects),
        "discovered_defects": list(discovered_defects),
        "parts":             list(client_parts),
        "labor":             list(labor),
        "total_labor_cost":  labor_cost,
        "billable_parts":    list(billable_parts),
        "total_parts_cost":  parts_cost,
        "grand_total":       labor_cost + parts_cost,
    }


def _run_combo(receipt_data: dict, tmp_path):
    """Run generation with mocked DB and return (worksheet, warnings)."""
    from openpyxl import load_workbook

    with (
        patch("src.services.excel_export.EXPORTS_DIR", tmp_path),
        patch("src.services.excel_export.get_tva", return_value=TVA),
        patch("src.services.excel_export.get_receipt_number", return_value=1),
        patch("src.services.excel_export.update_receipt_number"),
        patch("src.services.excel_export.get_defect_by_id",
              side_effect=lambda i: MOCK_DEFECTS.get(i)),
        patch("src.services.excel_export.get_part_by_id",
              side_effect=lambda i: MOCK_CLIENT_PARTS.get(i)),
        patch("src.services.excel_export.get_labor_by_id",
              side_effect=lambda i: MOCK_LABOR.get(i)),
    ):
        output_path, warnings = generate_receipt_excel(receipt_data)

    wb = load_workbook(output_path)
    ws = wb.active
    _assert_structure(ws, receipt_data)
    return ws, warnings


def _assert_structure(ws, receipt_data: dict):
    """Structural invariants that must hold for every generated receipt."""
    n = len(receipt_data.get("labor", []))
    m = len(receipt_data.get("billable_parts", []))
    d = len(receipt_data.get("defects", []))
    p = len(receipt_data.get("parts", []))
    n_disc = len(receipt_data.get("discovered_defects", []))
    row_offset   = max(5, d, p + 2) - 5
    disc_extra   = max(0, n_disc - 5)
    total_offset = row_offset + disc_extra

    manopera_row    = 32 + total_offset
    labor_base_row  = 36 + total_offset
    parts_start_row = 42 + max(0, n - 1) + total_offset

    assert ws[f"A{manopera_row}"].value == "MANOPERA", (
        f"A{manopera_row} trebuie sa fie 'MANOPERA', dar e: {ws[f'A{manopera_row}'].value!r}"
    )
    for col in "BCDEF":
        val = ws[f"{col}{manopera_row}"].value
        assert val in (None, ""), f"{col}{manopera_row} trebuie sa fie gol, dar e: {val!r}"

    total_labor_label_row = labor_base_row + max(1, n)
    assert ws[f"B{total_labor_label_row}"].value == "Total manopera", (
        f"B{total_labor_label_row} trebuie sa fie 'Total manopera', "
        f"dar e: {ws[f'B{total_labor_label_row}'].value!r}  (N={n})"
    )

    piese_row = total_labor_label_row + 1
    assert ws[f"A{piese_row}"].value == "PIESE", (
        f"A{piese_row} trebuie sa fie 'PIESE', "
        f"dar e: {ws[f'A{piese_row}'].value!r}  (N={n})"
    )

    total_parts_label_row = parts_start_row + max(1, m)
    assert ws[f"B{total_parts_label_row}"].value == "Total piese", (
        f"B{total_parts_label_row} trebuie sa fie 'Total piese', "
        f"dar e: {ws[f'B{total_parts_label_row}'].value!r}  (N={n}, M={m})"
    )


# ===========================================================================
# Parametrized helpers
# ===========================================================================

def _defect_ids(n: int) -> list:
    return ALL_DEFECT_IDS[:n]

def _client_part_ids(n: int) -> list:
    return ALL_CLIENT_PART_IDS[:n]

def _labor_ids(n: int) -> list:
    return ALL_LABOR_IDS[:n]

def _billable_parts(n: int) -> list:
    return ALL_BILLABLE_PARTS[:n]


# ===========================================================================
# 1. Vary DEFECTS (0 → 5), all others at maximum
# ===========================================================================

@pytest.mark.parametrize("n_defects", [0, 1, 2, 3, 4, 5])
def test_vary_defects(tmp_path, n_defects):
    """Defects list size varies 0-5; everything else is at maximum."""
    data = _make_receipt(defects=_defect_ids(n_defects))
    ws, warnings = _run_combo(data, tmp_path)

    for i, did in enumerate(_defect_ids(n_defects)):
        expected = MOCK_DEFECTS[did]["defect_name"]
        actual   = ws[f"A{14 + i}"].value
        assert actual == expected, (
            f"Defect #{i+1} (n={n_defects}): A{14+i} trebuia '{expected}', dar e {actual!r}"
        )


# ===========================================================================
# 2. Vary DISCOVERED DEFECTS (0 → 5), all others at maximum
# ===========================================================================

@pytest.mark.parametrize("n_disc", [0, 1, 2, 3, 4, 5])
def test_vary_discovered_defects(tmp_path, n_disc):
    """Discovered defects list size varies 0-5; everything else at maximum."""
    data = _make_receipt(discovered_defects=_defect_ids(n_disc))
    ws, warnings = _run_combo(data, tmp_path)

    for i, did in enumerate(_defect_ids(n_disc)):
        expected = MOCK_DEFECTS[did]["defect_name"]
        actual   = ws[f"A{20 + i}"].value
        assert actual == expected, (
            f"Defect descoperit #{i+1} (n={n_disc}): A{20+i} trebuia '{expected}', dar e {actual!r}"
        )


# ===========================================================================
# 3. Vary CLIENT PARTS (0 → 3), all others at maximum
# ===========================================================================

@pytest.mark.parametrize("n_cp", [0, 1, 2, 3])
def test_vary_client_parts(tmp_path, n_cp):
    """Client-supplied parts list size varies 0-3; everything else at maximum."""
    data = _make_receipt(client_parts=_client_part_ids(n_cp))
    ws, warnings = _run_combo(data, tmp_path)

    for i, pid in enumerate(_client_part_ids(n_cp)):
        expected = MOCK_CLIENT_PARTS[pid]["part_name"]
        actual   = ws[f"C{14 + i}"].value
        assert actual == expected, (
            f"Piesă client #{i+1} (n={n_cp}): C{14+i} trebuia '{expected}', dar e {actual!r}"
        )


# ===========================================================================
# 4. Vary LABOR (0 → 3), all others at maximum
# ===========================================================================

@pytest.mark.parametrize("n_labor", [0, 1, 2, 3])
def test_vary_labor(tmp_path, n_labor):
    """Labor list size varies 0-3; everything else at maximum."""
    labor_cost = 300.0 * n_labor
    data = _make_receipt(labor=_labor_ids(n_labor))
    data["total_labor_cost"] = labor_cost
    data["grand_total"] = labor_cost + data["total_parts_cost"]

    ws, warnings = _run_combo(data, tmp_path)

    for i, lid in enumerate(_labor_ids(n_labor)):
        expected = MOCK_LABOR[lid]["service_name"]
        actual   = ws[f"B{36 + i}"].value
        assert actual == expected, (
            f"Manoperă #{i+1} (n={n_labor}): B{36+i} trebuia '{expected}', dar e {actual!r}"
        )

    if n_labor > 0:
        total_row = 36 + n_labor
        assert ws[f"E{total_row}"].value == pytest.approx(labor_cost), (
            f"Total manoperă la E{total_row} incorect pentru n_labor={n_labor}"
        )


# ===========================================================================
# 5. Vary BILLABLE PARTS (0 → 3), all others at maximum
# ===========================================================================

@pytest.mark.parametrize("n_bp", [0, 1, 2, 3])
def test_vary_billable_parts(tmp_path, n_bp):
    """Billable parts list size varies 0-3; everything else at maximum."""
    bp = _billable_parts(n_bp)
    parts_cost = sum(p["units"] * p["price_per_unit"] for p in bp)
    n_labor = len(ALL_LABOR_IDS)          # 3
    labor_cost = 300.0 * n_labor

    data = _make_receipt(billable_parts=bp)
    data["total_parts_cost"] = parts_cost
    data["grand_total"]      = labor_cost + parts_cost

    ws, warnings = _run_combo(data, tmp_path)

    parts_start_row = 42 + max(0, n_labor - 1)  # = 44 for n_labor=3

    for i, part in enumerate(bp):
        row      = parts_start_row + i
        expected = part["part_name"]
        actual   = ws[f"B{row}"].value
        assert actual == expected, (
            f"Piesă facturabilă #{i+1} (n_bp={n_bp}): B{row} trebuia '{expected}', dar e {actual!r}"
        )
        assert ws[f"C{row}"].value == pytest.approx(part["units"]), (
            f"Cantitate piesă #{i+1} la C{row} incorectă (n_bp={n_bp})"
        )
        assert ws[f"D{row}"].value == pytest.approx(part["price_per_unit"]), (
            f"Preț/unitate piesă #{i+1} la D{row} incorect (n_bp={n_bp})"
        )
        assert ws[f"E{row}"].value == pytest.approx(part["units"] * part["price_per_unit"]), (
            f"Subtotal piesă #{i+1} la E{row} incorect (n_bp={n_bp})"
        )

    if n_bp > 0:
        total_row = parts_start_row + n_bp
        assert ws[f"E{total_row}"].value == pytest.approx(parts_cost), (
            f"Total piese la E{total_row} incorect pentru n_bp={n_bp}"
        )


# ===========================================================================
# 6. Pairwise combinations — every pair of section sizes (0 and max)
# ===========================================================================
#
# For each pair of sections we test all four corner combinations:
#   (0, 0), (0, max), (max, 0), (max, max)
# This catches interactions between sections (e.g. labor row-insert shifting
# the parts section when one or both are empty/full).
# ===========================================================================

_PAIR_PARAMS = [
    # (n_labor, n_bp)
    pytest.param(0, 0, id="labor=0_bp=0"),
    pytest.param(0, 3, id="labor=0_bp=3"),
    pytest.param(3, 0, id="labor=3_bp=0"),
    pytest.param(3, 3, id="labor=3_bp=3"),
    pytest.param(1, 0, id="labor=1_bp=0"),
    pytest.param(1, 1, id="labor=1_bp=1"),
    pytest.param(2, 2, id="labor=2_bp=2"),
]


@pytest.mark.parametrize("n_labor,n_bp", _PAIR_PARAMS)
def test_labor_and_parts_combinations(tmp_path, n_labor, n_bp):
    """Pairwise labor × billable_parts combinations — validates row placement."""
    bp         = _billable_parts(n_bp)
    parts_cost = sum(p["units"] * p["price_per_unit"] for p in bp)
    labor_cost = 300.0 * n_labor

    data = _make_receipt(labor=_labor_ids(n_labor), billable_parts=bp)
    data["total_labor_cost"] = labor_cost
    data["total_parts_cost"] = parts_cost
    data["grand_total"]      = labor_cost + parts_cost

    ws, _ = _run_combo(data, tmp_path)

    # ── Labor names ───────────────────────────────────────────────────────
    for i, lid in enumerate(_labor_ids(n_labor)):
        row      = 36 + i
        expected = MOCK_LABOR[lid]["service_name"]
        actual   = ws[f"B{row}"].value
        assert actual == expected, (
            f"Manoperă #{i+1}: B{row}='{actual}' (labor={n_labor}, bp={n_bp})"
        )

    # ── Labor total ───────────────────────────────────────────────────────
    if n_labor > 0:
        total_row = 36 + n_labor
        assert ws[f"E{total_row}"].value == pytest.approx(labor_cost), (
            f"E{total_row} labor total gresit (labor={n_labor}, bp={n_bp})"
        )

    # ── Billable parts ────────────────────────────────────────────────────
    parts_start_row = 42 + max(0, n_labor - 1)
    for i, part in enumerate(bp):
        row = parts_start_row + i
        assert ws[f"B{row}"].value == part["part_name"], (
            f"Piesă #{i+1}: B{row} (labor={n_labor}, bp={n_bp})"
        )
        assert ws[f"E{row}"].value == pytest.approx(part["units"] * part["price_per_unit"]), (
            f"Subtotal piesă #{i+1}: E{row} (labor={n_labor}, bp={n_bp})"
        )

    # ── Parts total ───────────────────────────────────────────────────────
    if n_bp > 0:
        total_row = parts_start_row + n_bp
        assert ws[f"E{total_row}"].value == pytest.approx(parts_cost), (
            f"E{total_row} parts total gresit (labor={n_labor}, bp={n_bp})"
        )

    # ── Grand total ───────────────────────────────────────────────────────
    if n_bp > 0:
        grand_row = parts_start_row + n_bp + 1
    elif n_labor > 0:
        grand_row = 43 + n_labor
    else:
        grand_row = 44

    assert ws[f"F{grand_row}"].value == pytest.approx(labor_cost + parts_cost), (
        f"F{grand_row} grand total gresit (labor={n_labor}, bp={n_bp})"
    )
