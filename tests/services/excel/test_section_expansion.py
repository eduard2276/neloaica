"""Tests for dynamic row expansion in the defects / client-parts section.

Template section layout (rows 13-18, baseline):

  Row 13  A="Defecte reclamate de client:"   C="Urmatoarele piese sunt furnizate de client."
  Row 14  A=<defect slot 1>                  C=<part slot 1>
  Row 15  A=<defect slot 2>                  C=<part slot 2>
  Row 16  A=<defect slot 3>                  C=<part slot 3>
  Row 17  A=<defect slot 4>                  C="Renunt la garantia acestor piese."
  Row 18  A=<defect slot 5>                  C="Semnatura client:"
  Row 19  next section ...

When N defects or M client parts exceed the template capacity (5 / 3),
extra rows are inserted at row 17 (just before the signature sub-section),
which shifts the signature rows downward automatically.

Row formula:
  section_rows = max(5, N_defects, N_parts + 2)
  extra_rows   = section_rows - 5
  section_end  = 13 + section_rows

Invariants that MUST hold regardless of N / M:
  • A13                    = "Defecte reclamate de client:"
  • A(14) .. A(13 + N)    = defect names in insertion order
  • C(14) .. C(13 + M)    = part names in insertion order
  • C(section_end - 1)    = "Renunt la garantia acestor piese."
  • C(section_end)        = "Semnatura client:"

Sections below the rectangle shift by extra_rows:
  • "MANOPERA" header moves from A32 → A(32 + extra_rows)
  • Discovered defects start at A(20 + extra_rows)
  • Labor base row shifts from 36 → 36 + extra_rows
"""

import pytest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from src.services.excel_export import generate_receipt_excel

# ---------------------------------------------------------------------------
# Mock data  (IDs 1-8 for defects, 10-15 for parts)
# ---------------------------------------------------------------------------

TVA = 21.0

MOCK_DEFECTS = {
    1: {"id": 1, "defect_name": "Defect 1"},
    2: {"id": 2, "defect_name": "Defect 2"},
    3: {"id": 3, "defect_name": "Defect 3"},
    4: {"id": 4, "defect_name": "Defect 4"},
    5: {"id": 5, "defect_name": "Defect 5"},
    6: {"id": 6, "defect_name": "Defect 6"},
    7: {"id": 7, "defect_name": "Defect 7"},
    8: {"id": 8, "defect_name": "Defect 8"},
}

MOCK_CLIENT_PARTS = {
    10: {"id": 10, "part_name": "Part 1"},
    11: {"id": 11, "part_name": "Part 2"},
    12: {"id": 12, "part_name": "Part 3"},
    13: {"id": 13, "part_name": "Part 4"},
    14: {"id": 14, "part_name": "Part 5"},
    15: {"id": 15, "part_name": "Part 6"},
}

MOCK_LABOR = {
    100: {"id": 100, "service_name": "Schimb ulei"},
    101: {"id": 101, "service_name": "Schimb filtre"},
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _section_end(n_defects: int, n_parts: int) -> int:
    """Row number of 'Semnatura client:' (last row of the expandable section)."""
    return 13 + max(5, n_defects, n_parts + 2)


def _extra_rows(n_defects: int, n_parts: int) -> int:
    return max(5, n_defects, n_parts + 2) - 5


def _run(receipt_data: dict, tmp_path: Path, receipt_number: int = 1):
    """Generate with all DB mocks and return (worksheet, warnings)."""
    with (
        patch("src.services.excel_export.EXPORTS_DIR", tmp_path),
        patch("src.services.excel_export.get_tva", return_value=TVA),
        patch("src.services.excel_export.get_receipt_number", return_value=receipt_number),
        patch("src.services.excel_export.update_receipt_number"),
        patch(
            "src.services.excel_export.get_defect_by_id",
            side_effect=lambda i: MOCK_DEFECTS.get(i),
        ),
        patch(
            "src.services.excel_export.get_part_by_id",
            side_effect=lambda i: MOCK_CLIENT_PARTS.get(i),
        ),
        patch(
            "src.services.excel_export.get_labor_by_id",
            side_effect=lambda i: MOCK_LABOR.get(i),
        ),
    ):
        output_path, warnings = generate_receipt_excel(receipt_data)

    wb = load_workbook(output_path)
    ws = wb.active
    return ws, warnings


def _base_receipt(**overrides) -> dict:
    base = {
        "client_name": "Test Client",
        "client_address": "",
        "model": "Test Car",
        "plate_number": "XX00TST",
        "vin": "",
        "kilometers": "0",
        "executant_name": "",
        "date": "07.05.2026",
        "estimate_cost": 0.0,
        "estimated_final_date": "",
        "defects": [],
        "discovered_defects": [],
        "parts": [],
        "labor": [],
        "total_labor_cost": 0.0,
        "billable_parts": [],
        "total_parts_cost": 0.0,
        "grand_total": 0.0,
    }
    base.update(overrides)
    return base


# ===========================================================================
# Section header anchor
# ===========================================================================

class TestSectionHeaderAnchor:

    def test_a13_is_section_header_with_no_data(self, tmp_path):
        """A13 must always be 'Defecte reclamate de client:' (baseline)."""
        ws, _ = _run(_base_receipt(), tmp_path)
        assert ws["A13"].value == "Defecte reclamate de client:", (
            f"A13: {ws['A13'].value!r}"
        )

    def test_a13_is_section_header_with_many_defects(self, tmp_path):
        """A13 stays put even when rows are inserted below it."""
        ws, _ = _run(_base_receipt(defects=list(range(1, 9))), tmp_path)
        assert ws["A13"].value == "Defecte reclamate de client:", (
            f"A13 changed after row insertion: {ws['A13'].value!r}"
        )

    def test_a13_is_section_header_with_many_parts(self, tmp_path):
        """A13 stays put when extra part rows are inserted."""
        ws, _ = _run(_base_receipt(parts=list(range(10, 15))), tmp_path)
        assert ws["A13"].value == "Defecte reclamate de client:", (
            f"A13 changed after row insertion: {ws['A13'].value!r}"
        )


# ===========================================================================
# Signature position — baseline (no expansion)
# ===========================================================================

class TestSignaturePositionBaseline:

    def test_renunt_at_c17_with_no_data(self, tmp_path):
        """With no defects / parts, 'Renunt la garantia...' stays at C17."""
        ws, _ = _run(_base_receipt(), tmp_path)
        assert ws["C17"].value == "Renunt la garantia acestor piese.", (
            f"C17: {ws['C17'].value!r}"
        )

    def test_semnatura_at_c18_with_no_data(self, tmp_path):
        """With no defects / parts, 'Semnatura client:' stays at C18."""
        ws, _ = _run(_base_receipt(), tmp_path)
        assert ws["C18"].value == "Semnatura client:", (
            f"C18: {ws['C18'].value!r}"
        )

    def test_signature_position_with_max_template_data(self, tmp_path):
        """5 defects + 3 parts (template max) → no expansion, C17/C18 unchanged."""
        ws, _ = _run(_base_receipt(defects=list(range(1, 6)), parts=list(range(10, 13))), tmp_path)
        assert ws["C17"].value == "Renunt la garantia acestor piese."
        assert ws["C18"].value == "Semnatura client:"


# ===========================================================================
# Bold separator border above 'Renunt la garantia' row
# ===========================================================================

class TestSeparatorBorderPosition:
    """The medium bottom border on C–F must always be just above 'Renunt la garantia'."""

    def _has_medium_bottom(self, cell) -> bool:
        b = cell.border
        return bool(b.bottom and b.bottom.style == "medium")

    def _border_row(self, n_defects: int, n_parts: int) -> int:
        """Expected row of the medium bottom border = row above 'Renunt'."""
        return _section_end(n_defects, n_parts) - 2  # section_end - 2 = 16 + extra_rows

    def test_border_at_row_16_no_data(self, tmp_path):
        """Baseline: medium bottom border on C16–F16."""
        ws, _ = _run(_base_receipt(), tmp_path)
        r = self._border_row(0, 0)   # 16
        for col in ["C", "D", "E", "F"]:
            assert self._has_medium_bottom(ws[f"{col}{r}"]), (
                f"{col}{r} should have medium bottom border (no expansion)"
            )

    def test_border_absent_from_row_16_after_expansion(self, tmp_path):
        """After expansion the medium bottom border must leave row 16."""
        ws, _ = _run(_base_receipt(defects=list(range(1, 9))), tmp_path)
        for col in ["C", "D", "E", "F"]:
            b = ws[f"{col}16"].border
            assert not (b.bottom and b.bottom.style == "medium"), (
                f"{col}16 still has medium bottom border after expansion"
            )

    def test_border_moves_with_8_defects(self, tmp_path):
        """8 defects (extra_rows=3) → border at row 19 (above 'Renunt' at 20)."""
        ws, _ = _run(_base_receipt(defects=list(range(1, 9))), tmp_path)
        r = self._border_row(8, 0)   # 19
        for col in ["C", "D", "E", "F"]:
            assert self._has_medium_bottom(ws[f"{col}{r}"]), (
                f"{col}{r} should have medium bottom border with 8 defects"
            )

    def test_border_moves_with_5_parts(self, tmp_path):
        """5 parts (extra_rows=2) → border at row 18 (above 'Renunt' at 19)."""
        ws, _ = _run(_base_receipt(parts=list(range(10, 15))), tmp_path)
        r = self._border_row(0, 5)   # 18
        for col in ["C", "D", "E", "F"]:
            assert self._has_medium_bottom(ws[f"{col}{r}"]), (
                f"{col}{r} should have medium bottom border with 5 parts"
            )

    def test_border_moves_with_8_defects_and_5_parts(self, tmp_path):
        """8 defects + 5 parts (extra_rows=3) → border at row 19."""
        ws, _ = _run(_base_receipt(defects=list(range(1, 9)), parts=list(range(10, 15))), tmp_path)
        r = self._border_row(8, 5)   # 19
        for col in ["C", "D", "E", "F"]:
            assert self._has_medium_bottom(ws[f"{col}{r}"]), (
                f"{col}{r} should have medium bottom border with 8 def + 5 parts"
            )


# ===========================================================================
# Many defects  (N > 5)
# ===========================================================================

class TestManyDefects:
    """Tests with 8 defects  →  section_rows = 8, extra_rows = 3, section_end = 21."""

    N = 8
    DEFECT_IDS = list(range(1, 9))   # [1..8]
    SE = _section_end(N, 0)          # = 21

    def test_all_defects_written(self, tmp_path):
        """All 8 defects must appear in A14..A21 with no truncation."""
        ws, warnings = _run(_base_receipt(defects=self.DEFECT_IDS), tmp_path)
        for i, did in enumerate(self.DEFECT_IDS):
            expected = MOCK_DEFECTS[did]["defect_name"]
            actual = ws[f"A{14 + i}"].value
            assert actual == expected, (
                f"Defect {i + 1}: A{14 + i} expected {expected!r}, got {actual!r}"
            )

    def test_no_overflow_warning_when_rows_expanded(self, tmp_path):
        """No defect-overflow warning when the section expands to fit."""
        _, warnings = _run(_base_receipt(defects=self.DEFECT_IDS), tmp_path)
        assert not any("defect" in w.lower() for w in warnings), (
            f"Unexpected defect overflow warning: {warnings}"
        )

    def test_renunt_shifts_to_correct_row(self, tmp_path):
        """'Renunt la garantia...' must be at C(section_end - 1)."""
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS), tmp_path)
        row = self.SE - 1   # C20
        assert ws[f"C{row}"].value == "Renunt la garantia acestor piese.", (
            f"C{row}: {ws[f'C{row}'].value!r}  (section_end={self.SE})"
        )

    def test_semnatura_shifts_to_correct_row(self, tmp_path):
        """'Semnatura client:' must be at C(section_end)."""
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS), tmp_path)
        assert ws[f"C{self.SE}"].value == "Semnatura client:", (
            f"C{self.SE}: {ws[f'C{self.SE}'].value!r}"
        )

    def test_manopera_shifts_down(self, tmp_path):
        """'MANOPERA' header moves from A32 to A(32 + extra_rows)."""
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS), tmp_path)
        er = _extra_rows(self.N, 0)            # 3
        expected_row = 32 + er                 # 35
        assert ws[f"A{expected_row}"].value == "MANOPERA", (
            f"A{expected_row}: {ws[f'A{expected_row}'].value!r}  (extra_rows={er})"
        )

    def test_discovered_defects_shift_down(self, tmp_path):
        """Discovered defects start at A(20 + extra_rows) after expansion."""
        disc = [1, 2]
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS, discovered_defects=disc), tmp_path)
        er = _extra_rows(self.N, 0)
        for i, did in enumerate(disc):
            expected = MOCK_DEFECTS[did]["defect_name"]
            row = 20 + er + i
            assert ws[f"A{row}"].value == expected, (
                f"Discovered defect {i + 1}: A{row} expected {expected!r}, "
                f"got {ws[f'A{row}'].value!r}"
            )


# ===========================================================================
# Many parts  (M > 3)
# ===========================================================================

class TestManyParts:
    """Tests with 5 parts  →  section_rows = 7, extra_rows = 2, section_end = 20."""

    M = 5
    PART_IDS = list(range(10, 15))   # [10..14]
    SE = _section_end(0, M)          # = 20

    def test_all_parts_written(self, tmp_path):
        """All 5 parts must appear in C14..C18 with no truncation."""
        ws, warnings = _run(_base_receipt(parts=self.PART_IDS), tmp_path)
        for i, pid in enumerate(self.PART_IDS):
            expected = MOCK_CLIENT_PARTS[pid]["part_name"]
            actual = ws[f"C{14 + i}"].value
            assert actual == expected, (
                f"Part {i + 1}: C{14 + i} expected {expected!r}, got {actual!r}"
            )

    def test_no_overflow_warning_when_rows_expanded(self, tmp_path):
        """No parts-overflow warning when the section expands to fit."""
        _, warnings = _run(_base_receipt(parts=self.PART_IDS), tmp_path)
        assert not any("parts received" in w.lower() for w in warnings), (
            f"Unexpected parts overflow warning: {warnings}"
        )

    def test_renunt_shifts_to_correct_row(self, tmp_path):
        """'Renunt la garantia...' must be at C(section_end - 1)."""
        ws, _ = _run(_base_receipt(parts=self.PART_IDS), tmp_path)
        row = self.SE - 1   # C19
        assert ws[f"C{row}"].value == "Renunt la garantia acestor piese.", (
            f"C{row}: {ws[f'C{row}'].value!r}  (section_end={self.SE})"
        )

    def test_semnatura_shifts_to_correct_row(self, tmp_path):
        """'Semnatura client:' must be at C(section_end)."""
        ws, _ = _run(_base_receipt(parts=self.PART_IDS), tmp_path)
        assert ws[f"C{self.SE}"].value == "Semnatura client:", (
            f"C{self.SE}: {ws[f'C{self.SE}'].value!r}"
        )

    def test_manopera_shifts_down(self, tmp_path):
        """'MANOPERA' header moves from A32 to A(32 + extra_rows)."""
        ws, _ = _run(_base_receipt(parts=self.PART_IDS), tmp_path)
        er = _extra_rows(0, self.M)            # 2
        expected_row = 32 + er                 # 34
        assert ws[f"A{expected_row}"].value == "MANOPERA", (
            f"A{expected_row}: {ws[f'A{expected_row}'].value!r}  (extra_rows={er})"
        )


# ===========================================================================
# Many defects AND many parts  (N=8, M=5)
# ===========================================================================

class TestManyDefectsAndParts:
    """Tests with 8 defects + 5 parts  →  section_rows=8, extra_rows=3, section_end=21."""

    N = 8
    M = 5
    DEFECT_IDS = list(range(1, 9))
    PART_IDS   = list(range(10, 15))
    SE = _section_end(N, M)   # = 21

    def test_all_defects_written(self, tmp_path):
        """All 8 defects written to A14..A21."""
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS, parts=self.PART_IDS), tmp_path)
        for i, did in enumerate(self.DEFECT_IDS):
            expected = MOCK_DEFECTS[did]["defect_name"]
            assert ws[f"A{14 + i}"].value == expected, (
                f"Defect {i + 1}: A{14 + i} expected {expected!r}, "
                f"got {ws[f'A{14 + i}'].value!r}"
            )

    def test_all_parts_written(self, tmp_path):
        """All 5 parts written to C14..C18."""
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS, parts=self.PART_IDS), tmp_path)
        for i, pid in enumerate(self.PART_IDS):
            expected = MOCK_CLIENT_PARTS[pid]["part_name"]
            assert ws[f"C{14 + i}"].value == expected, (
                f"Part {i + 1}: C{14 + i} expected {expected!r}, "
                f"got {ws[f'C{14 + i}'].value!r}"
            )

    def test_renunt_at_correct_row(self, tmp_path):
        """'Renunt la garantia...' at C(section_end - 1) = C20."""
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS, parts=self.PART_IDS), tmp_path)
        row = self.SE - 1
        assert ws[f"C{row}"].value == "Renunt la garantia acestor piese.", (
            f"C{row}: {ws[f'C{row}'].value!r}"
        )

    def test_semnatura_at_correct_row(self, tmp_path):
        """'Semnatura client:' at C(section_end) = C21."""
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS, parts=self.PART_IDS), tmp_path)
        assert ws[f"C{self.SE}"].value == "Semnatura client:", (
            f"C{self.SE}: {ws[f'C{self.SE}'].value!r}"
        )

    def test_manopera_shifts_down(self, tmp_path):
        """'MANOPERA' at A(32 + extra_rows) = A35."""
        ws, _ = _run(_base_receipt(defects=self.DEFECT_IDS, parts=self.PART_IDS), tmp_path)
        er = _extra_rows(self.N, self.M)  # 3
        expected_row = 32 + er
        assert ws[f"A{expected_row}"].value == "MANOPERA", (
            f"A{expected_row}: {ws[f'A{expected_row}'].value!r}"
        )

    def test_no_overflow_warnings(self, tmp_path):
        """No overflow warnings when both sections are expanded."""
        _, warnings = _run(_base_receipt(defects=self.DEFECT_IDS, parts=self.PART_IDS), tmp_path)
        assert warnings == [], f"Unexpected warnings: {warnings}"

    def test_labor_still_written_correctly(self, tmp_path):
        """Labor must still be written at the shifted row (36 + extra_rows)."""
        er = _extra_rows(self.N, self.M)          # 3
        labor_row = 36 + er                       # 39
        part = {"part_id": None, "part_name": "Filtru ulei", "units": 1.0, "price_per_unit": 50.0}
        ws, _ = _run(
            _base_receipt(
                defects=self.DEFECT_IDS,
                parts=self.PART_IDS,
                labor=[100],
                total_labor_cost=300.0,
                billable_parts=[part],
                total_parts_cost=50.0,
            ),
            tmp_path,
        )
        assert ws[f"B{labor_row}"].value == "Schimb ulei", (
            f"Labor at B{labor_row}: {ws[f'B{labor_row}'].value!r}  (extra_rows={er})"
        )


# ===========================================================================
# Full receipt with expansion: verify all downstream sections
# ===========================================================================

class TestFullReceiptWithExpansion:
    """
    8 defects + 5 client parts  →  section_rows=8, extra_rows=3, row_offset=3.
    Receipt also has 2 labor items and 3 billable parts.

    Expected row layout after expansion:
      MANOPERA header        : A35  (32 + 3)
      Labor item 1           : B39  (36 + 3)
      Labor item 2           : B40  (row inserted)
      Total manopera         : E41  (39 + 2)
      PIESE header           : A42  (41 + 1)
      Billable part 1        : B46  (42 + 1 labor-insert + 3 row_offset)
      Billable part 2        : B47  (row inserted)
      Billable part 3        : B48  (row inserted)
      Total piese            : E49  (46 + 3)
      Grand total            : F50  (49 + 1)
      Executant              : B56  (50 + 1 extra-labor + 2 extra-parts + 3 row_offset)
    """

    N_DEF  = 8
    N_PART = 5
    DEFECT_IDS      = list(range(1, 9))    # IDs 1-8
    CLIENT_PART_IDS = list(range(10, 15))  # IDs 10-14
    LABOR_IDS       = [100, 101]
    BILLABLE_PARTS  = [
        {"part_name": "Ulei motor",  "units": 1.0, "price_per_unit": 60.0},
        {"part_name": "Filtru ulei", "units": 1.0, "price_per_unit": 30.0},
        {"part_name": "Bujii",       "units": 4.0, "price_per_unit": 15.0},
    ]
    TOTAL_LABOR = 500.0
    TOTAL_PARTS = 150.0

    # ── pre-computed row constants ─────────────────────────────────────────
    # ER = max(5, 8, 5+2) - 5 = 3
    ER              = 3
    LABOR_BASE      = 39   # 36 + 3
    TOTAL_LABOR_ROW = 41   # 39 + 2 (len LABOR_IDS)
    PARTS_START     = 46   # 42 + max(0, 2-1) + 3
    TOTAL_PARTS_ROW = 49   # 46 + 3 (len BILLABLE_PARTS)
    GRAND_TOTAL_ROW = 50   # 49 + 1
    EXECUTANT_ROW   = 56   # 50 + 1 extra-labor + 2 extra-parts + 3 row_offset

    def _receipt(self, **extra):
        data = _base_receipt(
            defects=self.DEFECT_IDS,
            parts=self.CLIENT_PART_IDS,
            labor=self.LABOR_IDS,
            total_labor_cost=self.TOTAL_LABOR,
            billable_parts=self.BILLABLE_PARTS,
            total_parts_cost=self.TOTAL_PARTS,
            grand_total=self.TOTAL_LABOR + self.TOTAL_PARTS,
            executant_name="Ion Marius",
        )
        data.update(extra)
        return data

    # ── labor ──────────────────────────────────────────────────────────────

    def test_first_labor_written_to_correct_row(self, tmp_path):
        """First labor service at B(LABOR_BASE) = B39 after row_offset=3."""
        ws, _ = _run(self._receipt(), tmp_path)
        assert ws[f"B{self.LABOR_BASE}"].value == "Schimb ulei", (
            f"B{self.LABOR_BASE}: {ws[f'B{self.LABOR_BASE}'].value!r}"
        )

    def test_second_labor_written_to_correct_row(self, tmp_path):
        """Second labor service inserted at B(LABOR_BASE + 1) = B40."""
        ws, _ = _run(self._receipt(), tmp_path)
        row = self.LABOR_BASE + 1
        assert ws[f"B{row}"].value == "Schimb filtre", (
            f"B{row}: {ws[f'B{row}'].value!r}"
        )

    def test_total_labor_value_at_correct_row(self, tmp_path):
        """Total manopera value written to E(TOTAL_LABOR_ROW) = E41."""
        ws, _ = _run(self._receipt(), tmp_path)
        assert ws[f"E{self.TOTAL_LABOR_ROW}"].value == self.TOTAL_LABOR, (
            f"E{self.TOTAL_LABOR_ROW}: {ws[f'E{self.TOTAL_LABOR_ROW}'].value!r}  "
            f"expected {self.TOTAL_LABOR}"
        )

    def test_total_labor_tva_at_correct_row(self, tmp_path):
        """Labor TVA written to F(TOTAL_LABOR_ROW) = F41."""
        ws, _ = _run(self._receipt(), tmp_path)
        expected_tva = (self.TOTAL_LABOR * TVA) / (100 + TVA)
        actual = ws[f"F{self.TOTAL_LABOR_ROW}"].value
        assert actual is not None, f"F{self.TOTAL_LABOR_ROW} is None"
        assert abs(actual - expected_tva) < 0.01, (
            f"Labor TVA at F{self.TOTAL_LABOR_ROW}: got {actual}, expected ~{expected_tva:.2f}"
        )

    # ── billable parts ─────────────────────────────────────────────────────

    def test_billable_parts_names_at_correct_rows(self, tmp_path):
        """All 3 billable parts land at B46, B47, B48 (PARTS_START + i)."""
        ws, _ = _run(self._receipt(), tmp_path)
        for i, bp in enumerate(self.BILLABLE_PARTS):
            row = self.PARTS_START + i
            assert ws[f"B{row}"].value == bp["part_name"], (
                f"Billable part {i + 1}: B{row} expected {bp['part_name']!r}, "
                f"got {ws[f'B{row}'].value!r}"
            )

    def test_billable_part_units_at_correct_rows(self, tmp_path):
        """Part units written to column C at each part row."""
        ws, _ = _run(self._receipt(), tmp_path)
        for i, bp in enumerate(self.BILLABLE_PARTS):
            row = self.PARTS_START + i
            assert ws[f"C{row}"].value == bp["units"], (
                f"Part {i + 1} units: C{row} expected {bp['units']}, "
                f"got {ws[f'C{row}'].value!r}"
            )

    def test_billable_part_price_at_correct_rows(self, tmp_path):
        """Part price_per_unit written to column D at each part row."""
        ws, _ = _run(self._receipt(), tmp_path)
        for i, bp in enumerate(self.BILLABLE_PARTS):
            row = self.PARTS_START + i
            assert ws[f"D{row}"].value == bp["price_per_unit"], (
                f"Part {i + 1} price: D{row} expected {bp['price_per_unit']}, "
                f"got {ws[f'D{row}'].value!r}"
            )

    def test_billable_part_subtotals_at_correct_rows(self, tmp_path):
        """Part subtotal (units × price) written to column E at each part row."""
        ws, _ = _run(self._receipt(), tmp_path)
        for i, bp in enumerate(self.BILLABLE_PARTS):
            row = self.PARTS_START + i
            expected = bp["units"] * bp["price_per_unit"]
            assert ws[f"E{row}"].value == expected, (
                f"Part {i + 1} subtotal: E{row} expected {expected}, "
                f"got {ws[f'E{row}'].value!r}"
            )

    def test_total_parts_value_at_correct_row(self, tmp_path):
        """Total piese value at E(TOTAL_PARTS_ROW) = E49."""
        ws, _ = _run(self._receipt(), tmp_path)
        assert ws[f"E{self.TOTAL_PARTS_ROW}"].value == self.TOTAL_PARTS, (
            f"E{self.TOTAL_PARTS_ROW}: {ws[f'E{self.TOTAL_PARTS_ROW}'].value!r}  "
            f"expected {self.TOTAL_PARTS}"
        )

    # ── grand total & executant ────────────────────────────────────────────

    def test_grand_total_at_correct_row(self, tmp_path):
        """Grand total written to F(GRAND_TOTAL_ROW) = F50."""
        ws, _ = _run(self._receipt(), tmp_path)
        expected = self.TOTAL_LABOR + self.TOTAL_PARTS
        assert ws[f"F{self.GRAND_TOTAL_ROW}"].value == expected, (
            f"F{self.GRAND_TOTAL_ROW}: {ws[f'F{self.GRAND_TOTAL_ROW}'].value!r}  "
            f"expected {expected}"
        )

    def test_executant_at_correct_row(self, tmp_path):
        """Executant name written to B(EXECUTANT_ROW) = B56."""
        ws, _ = _run(self._receipt(), tmp_path)
        assert ws[f"B{self.EXECUTANT_ROW}"].value == "Ion Marius", (
            f"B{self.EXECUTANT_ROW}: {ws[f'B{self.EXECUTANT_ROW}'].value!r}"
        )

    def test_no_warnings_on_full_receipt(self, tmp_path):
        """No warnings on a fully populated receipt with section expansion."""
        _, warnings = _run(self._receipt(), tmp_path)
        assert warnings == [], f"Unexpected warnings: {warnings}"
