"""Tests for receipt Excel generation (src/services/excel_export.py).

Each test generates a real .xlsx file into a temp directory (pytest's
tmp_path fixture) and reads it back with openpyxl to assert cell values.
All database calls are mocked so the tests have no database dependency.

Row-layout reference (template base, zero inserts):
  B36          – first labor service
  E(36+N)      – labor total         (N = number of labor items)
  B42          – first billable part  (shifts by N-1 extra labor inserts)
  E(42+N+M-1)  – parts total         (M = number of billable parts)
  F(43+N+M-1)  – grand total
  B50          – executant name       (shifts by (N-1)+(M-1) extra inserts)
"""

from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from src.services.excel_export import generate_receipt_excel, template_exists

# ---------------------------------------------------------------------------
# Test data
# ---------------------------------------------------------------------------

TVA = 21.0

MOCK_DEFECTS = {
    1: {"id": 1, "defect_name": "Placute frana uzate"},
    2: {"id": 2, "defect_name": "Scurgere ulei motor"},
    3: {"id": 3, "defect_name": "Presiune scazuta anvelope"},
    4: {"id": 4, "defect_name": "Nivel lichid racire scazut"},
    5: {"id": 5, "defect_name": "Stergator parbriz uzat"},
    6: {"id": 6, "defect_name": "Bec far ars"},
}

MOCK_CLIENT_PARTS = {
    10: {"id": 10, "part_name": "Filtru ulei"},
    11: {"id": 11, "part_name": "Filtru aer"},
    12: {"id": 12, "part_name": "Filtru habitaclu"},
    13: {"id": 13, "part_name": "Filtru combustibil"},
}

MOCK_LABOR = {
    100: {"id": 100, "service_name": "Schimb ulei"},
    101: {"id": 101, "service_name": "Inlocuire placute frana"},
    102: {"id": 102, "service_name": "Geometrie roti"},
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _base_receipt(**overrides) -> dict:
    """Return a minimal valid receipt_data dict."""
    base = {
        "client_name": "Ion Popescu",
        "client_address": "Str. Principala 1",
        "model": "Dacia Logan",
        "plate_number": "B123ABC",
        "vin": "VF1AAAAA000000001",
        "kilometers": "150000",
        "executant_name": "Mihai Ionescu",
        "date": "07.05.2026",
        "estimate_cost": 500.0,
        "estimated_final_date": "10.05.2026",
        "defects": [],
        "discovered_defects": [],
        "parts": [],
        "labor": [],
        "total_labor_cost": 0.0,
        "billable_parts": [],
        "total_parts_cost": 0.0,
        "grand_total": 0.0,
    }
    base.update(overrides)
    return base


def _run(receipt_data: dict, tmp_exports: Path, receipt_number: int = 1):
    """
    Run generate_receipt_excel with all external (DB) calls mocked.

    Returns (worksheet, warning_messages) after opening the generated file.
    """
    with (
        patch("src.services.excel_export.EXPORTS_DIR", tmp_exports),
        patch("src.services.excel_export.get_tva", return_value=TVA),
        patch("src.services.excel_export.get_receipt_number", return_value=receipt_number),
        patch("src.services.excel_export.update_receipt_number"),
        patch(
            "src.services.excel_export.get_defect_by_id",
            side_effect=lambda i: MOCK_DEFECTS.get(i),
        ),
        patch(
            "src.services.excel_export.get_part_by_id",
            side_effect=lambda i: MOCK_CLIENT_PARTS.get(i),
        ),
        patch(
            "src.services.excel_export.get_labor_by_id",
            side_effect=lambda i: MOCK_LABOR.get(i),
        ),
    ):
        output_path, warnings = generate_receipt_excel(receipt_data)

    wb = load_workbook(output_path)
    ws = wb.active
    _assert_structure(ws, receipt_data)
    return ws, warnings


def _assert_structure(ws, receipt_data: dict):
    """
    Assert structural invariants that must hold for every generated receipt.

    Row formulas (N = labor, M = billable_parts, D = defects, P = client_parts):
      row_offset      = max(5, D, P+2) - 5  (extra rows from section expansion)
      MANOPERA row    = 32 + row_offset
      Total manopera  = (36 + row_offset) + max(1, N)
      PIESE row       = Total manopera + 1
      parts_start_row = 42 + max(0, N-1) + row_offset
      Total piese     = parts_start_row + max(1, M)
    """
    n = len(receipt_data.get("labor", []))
    m = len(receipt_data.get("billable_parts", []))
    d = len(receipt_data.get("defects", []))
    p = len(receipt_data.get("parts", []))
    n_disc = len(receipt_data.get("discovered_defects", []))
    row_offset = max(5, d, p + 2) - 5
    disc_extra = max(0, n_disc - 5)
    total_offset = row_offset + disc_extra

    manopera_row = 32 + total_offset
    labor_base_row = 36 + total_offset
    parts_start_row = 42 + max(0, n - 1) + total_offset

    # ── MANOPERA header ───────────────────────────────────────────────────
    assert (
        ws[f"A{manopera_row}"].value == "MANOPERA"
    ), f"A{manopera_row} trebuie sa fie 'MANOPERA', dar e: {ws[f'A{manopera_row}'].value!r}"
    for col in "BCDEF":
        val = ws[f"{col}{manopera_row}"].value
        assert val in (None, ""), f"{col}{manopera_row} trebuie sa fie gol, dar e: {val!r}"

    # ── "Total manopera" one row after last labor entry ───────────────────
    total_labor_label_row = labor_base_row + max(1, n)
    assert ws[f"B{total_labor_label_row}"].value == "Total manopera", (
        f"B{total_labor_label_row} trebuie sa fie 'Total manopera', "
        f"dar e: {ws[f'B{total_labor_label_row}'].value!r}"
    )

    # ── "PIESE" one row below "Total manopera" ───────────────────────────
    piese_row = total_labor_label_row + 1
    assert ws[f"A{piese_row}"].value == "PIESE", (
        f"A{piese_row} trebuie sa fie 'PIESE', " f"dar e: {ws[f'A{piese_row}'].value!r}"
    )

    # ── "Total piese" one row after last billable part ────────────────────
    total_parts_label_row = parts_start_row + max(1, m)
    assert ws[f"B{total_parts_label_row}"].value == "Total piese", (
        f"B{total_parts_label_row} trebuie sa fie 'Total piese', "
        f"dar e: {ws[f'B{total_parts_label_row}'].value!r}"
    )


# ---------------------------------------------------------------------------
# Prerequisites
# ---------------------------------------------------------------------------


def test_template_exists():
    """The Excel template must be present for any generation to work."""
    assert template_exists(), (
        "Template file 'Template-deviz.xlsx' is missing – "
        "all generation tests will fail without it."
    )


# ---------------------------------------------------------------------------
# Output format
# ---------------------------------------------------------------------------


def test_returns_path_and_warning_list(tmp_path):
    output_path, warnings = (
        _run.__wrapped__ if hasattr(_run, "__wrapped__") else _run(_base_receipt(), tmp_path)
    )
    # _run returns (ws, warnings); the raw function returns (str_path, list)
    with (
        patch("src.services.excel_export.EXPORTS_DIR", tmp_path),
        patch("src.services.excel_export.get_tva", return_value=TVA),
        patch("src.services.excel_export.get_receipt_number", return_value=1),
        patch("src.services.excel_export.update_receipt_number"),
        patch("src.services.excel_export.get_defect_by_id", return_value=None),
        patch("src.services.excel_export.get_part_by_id", return_value=None),
        patch("src.services.excel_export.get_labor_by_id", return_value=None),
    ):
        result = generate_receipt_excel(_base_receipt())

    assert isinstance(result, tuple) and len(result) == 2
    path_str, warnings = result
    assert isinstance(path_str, str)
    assert isinstance(warnings, list)
    assert Path(path_str).exists()


def test_output_file_is_readable_xlsx(tmp_path):
    ws, _ = _run(_base_receipt(), tmp_path)
    assert ws is not None


# ---------------------------------------------------------------------------
# Client / car fields
# ---------------------------------------------------------------------------


def test_client_name_in_b10(tmp_path):
    ws, _ = _run(_base_receipt(client_name="Maria Ionescu"), tmp_path)
    assert ws["B10"].value == "Maria Ionescu"


def test_client_address_in_b12(tmp_path):
    ws, _ = _run(_base_receipt(client_address="Bd. Unirii 5"), tmp_path)
    assert ws["B12"].value == "Bd. Unirii 5"


def test_car_model_in_d10(tmp_path):
    ws, _ = _run(_base_receipt(model="VW Golf"), tmp_path)
    assert ws["D10"].value == "VW Golf"


def test_plate_number_in_e11(tmp_path):
    ws, _ = _run(_base_receipt(plate_number="CJ42XYZ"), tmp_path)
    assert ws["E11"].value == "CJ42XYZ"


def test_vin_in_e12(tmp_path):
    ws, _ = _run(_base_receipt(vin="WVWZZZ1JZ3W386752"), tmp_path)
    assert ws["E12"].value == "WVWZZZ1JZ3W386752"


def test_kilometers_written_as_integer(tmp_path):
    ws, _ = _run(_base_receipt(kilometers="75000"), tmp_path)
    assert ws["F10"].value == 75000


def test_date_and_receipt_number_in_b8(tmp_path):
    ws, _ = _run(_base_receipt(date="07.05.2026"), tmp_path, receipt_number=7)
    cell = str(ws["B8"].value)
    assert "07.05.2026" in cell
    assert "7" in cell


# ---------------------------------------------------------------------------
# Receipt number increment
# ---------------------------------------------------------------------------


def test_receipt_number_incremented_after_generation(tmp_path):
    with (
        patch("src.services.excel_export.EXPORTS_DIR", tmp_path),
        patch("src.services.excel_export.get_tva", return_value=TVA),
        patch("src.services.excel_export.get_receipt_number", return_value=5),
        patch("src.services.excel_export.update_receipt_number") as mock_update,
        patch("src.services.excel_export.get_defect_by_id", return_value=None),
        patch("src.services.excel_export.get_part_by_id", return_value=None),
        patch("src.services.excel_export.get_labor_by_id", return_value=None),
    ):
        generate_receipt_excel(_base_receipt())
        mock_update.assert_called_once_with(6)


# ---------------------------------------------------------------------------
# Defects
# ---------------------------------------------------------------------------


def test_defects_written_to_a14_a16(tmp_path):
    ws, _ = _run(_base_receipt(defects=[1, 2, 3]), tmp_path)
    assert ws["A14"].value == "Placute frana uzate"
    assert ws["A15"].value == "Scurgere ulei motor"
    assert ws["A16"].value == "Presiune scazuta anvelope"


def test_defects_beyond_5_are_all_written(tmp_path):
    """A 6th defect is now written because the section expands (no hard limit)."""
    ws, _ = _run(_base_receipt(defects=[1, 2, 3, 4, 5, 6]), tmp_path)
    assert ws["A18"].value == "Stergator parbriz uzat"
    assert (
        ws["A19"].value == "Bec far ars"
    ), f"A19 should contain the 6th defect now that rows expand; got: {ws['A19'].value!r}"


def test_defects_no_overflow_warning_when_expanded(tmp_path):
    """No warning is emitted when the section expands to fit extra defects."""
    _, warnings = _run(_base_receipt(defects=[1, 2, 3, 4, 5, 6]), tmp_path)
    assert not any(
        "defect" in w.lower() for w in warnings
    ), f"Unexpected defect-overflow warning: {warnings}"


def test_discovered_defects_written_to_a20_a21(tmp_path):
    ws, _ = _run(_base_receipt(discovered_defects=[1, 2]), tmp_path)
    assert ws["A20"].value == "Placute frana uzate"
    assert ws["A21"].value == "Scurgere ulei motor"


def test_discovered_defects_beyond_5_are_all_written(tmp_path):
    """A 6th discovered defect is now written because the section expands."""
    ws, _ = _run(_base_receipt(discovered_defects=[1, 2, 3, 4, 5, 6]), tmp_path)
    assert ws["A24"].value == "Stergator parbriz uzat"
    assert (
        ws["A25"].value == "Bec far ars"
    ), f"A25 should contain the 6th discovered defect; got: {ws['A25'].value!r}"


def test_discovered_defects_no_overflow_warning_when_expanded(tmp_path):
    """No warning is emitted when the discovered section expands to fit."""
    _, warnings = _run(_base_receipt(discovered_defects=[1, 2, 3, 4, 5, 6]), tmp_path)
    assert not any(
        "defect" in w.lower() for w in warnings
    ), f"Unexpected discovered-defect-overflow warning: {warnings}"


# ---------------------------------------------------------------------------
# Client-supplied parts (C14-C16, max 3)
# ---------------------------------------------------------------------------


def test_client_parts_written_to_c14_c16(tmp_path):
    ws, _ = _run(_base_receipt(parts=[10, 11, 12]), tmp_path)
    assert ws["C14"].value == "Filtru ulei"
    assert ws["C15"].value == "Filtru aer"
    assert ws["C16"].value == "Filtru habitaclu"


def test_client_parts_beyond_3_are_all_written(tmp_path):
    """A 4th client part is now written because the section expands (no hard limit)."""
    ws, _ = _run(_base_receipt(parts=[10, 11, 12, 13]), tmp_path)
    assert ws["C14"].value == "Filtru ulei"
    assert ws["C15"].value == "Filtru aer"
    assert ws["C16"].value == "Filtru habitaclu"
    assert (
        ws["C17"].value == "Filtru combustibil"
    ), f"C17 should contain the 4th part now that rows expand; got: {ws['C17'].value!r}"


def test_client_parts_no_overflow_warning_when_expanded(tmp_path):
    """No warning is emitted when the section expands to fit extra parts."""
    _, warnings = _run(_base_receipt(parts=[10, 11, 12, 13]), tmp_path)
    assert not any(
        "parts received" in w.lower() for w in warnings
    ), f"Unexpected parts-overflow warning: {warnings}"


# ---------------------------------------------------------------------------
# Labor section
# ---------------------------------------------------------------------------


def test_single_labor_written_to_b36(tmp_path):
    ws, _ = _run(_base_receipt(labor=[100], total_labor_cost=300.0), tmp_path)
    assert (
        ws["B34"].value == "Operatie"
    ), "Header 'Operatie' trebuie sa fie la B34 (2 randuri deasupra B36)"
    assert ws["B36"].value == "Schimb ulei"


def test_labor_total_in_e37_for_1_item(tmp_path):
    """1 labor item → no row insert → total goes in E37 (36+1)."""
    ws, _ = _run(_base_receipt(labor=[100], total_labor_cost=350.0), tmp_path)
    assert ws["E37"].value == pytest.approx(350.0)


def test_labor_total_in_e39_for_3_items(tmp_path):
    """3 labor items (2 inserts) → total goes in E39 (36+3)."""
    ws, _ = _run(_base_receipt(labor=[100, 101, 102], total_labor_cost=900.0), tmp_path)
    assert ws["E39"].value == pytest.approx(900.0)


def test_labor_tva_calculated_correctly(tmp_path):
    """TVA in F37 = (total × TVA%) / (100 + TVA%)."""
    total = 500.0
    expected_tva = (total * TVA) / (100 + TVA)
    ws, _ = _run(_base_receipt(labor=[100], total_labor_cost=total), tmp_path)
    assert ws["F37"].value == pytest.approx(expected_tva, rel=1e-6)


def test_multiple_labor_names_written_sequentially(tmp_path):
    ws, _ = _run(_base_receipt(labor=[100, 101, 102], total_labor_cost=0.0), tmp_path)
    assert ws["B36"].value == "Schimb ulei"
    assert ws["B37"].value == "Inlocuire placute frana"
    assert ws["B38"].value == "Geometrie roti"


# ---------------------------------------------------------------------------
# Billable parts section
# ---------------------------------------------------------------------------


def test_billable_part_columns_1_labor_1_part(tmp_path):
    """1 labor + 1 billable part: part written at row 42 (parts_start_row=42)."""
    part = {"part_id": 10, "part_name": "Filtru ulei", "units": 2.0, "price_per_unit": 50.0}
    ws, _ = _run(
        _base_receipt(
            labor=[100],
            total_labor_cost=300.0,
            billable_parts=[part],
            total_parts_cost=100.0,
        ),
        tmp_path,
    )
    # parts_start_row=42 → header la B40 (2 randuri mai sus)
    assert (
        ws["B40"].value == "Piese auto-materiale"
    ), "Header 'Piese auto-materiale' trebuie sa fie la B40 cand parts_start_row=42"
    assert ws["B42"].value == "Filtru ulei"
    assert ws["C42"].value == pytest.approx(2.0)
    assert ws["D42"].value == pytest.approx(50.0)
    assert ws["E42"].value == pytest.approx(100.0)  # 2 × 50


def test_billable_part_row_no_labor(tmp_path):
    """0 labor + 1 billable part: part must start at row 42.

    No rows are inserted when labor is absent, so the template layout is
    unchanged: 'Piese auto-materiale' header at B40, first data row at B42.
    """
    part = {"part_id": 10, "part_name": "Filtru ulei", "units": 1.0, "price_per_unit": 50.0}
    ws, _ = _run(
        _base_receipt(
            billable_parts=[part],
            total_parts_cost=50.0,
        ),
        tmp_path,
    )
    # parts_start_row=42 → header la B40 (2 randuri mai sus)
    assert (
        ws["B40"].value == "Piese auto-materiale"
    ), "Header 'Piese auto-materiale' trebuie sa fie la B40 cand nu exista manopera"
    assert (
        ws["B42"].value == "Filtru ulei"
    ), "Piesa trebuie sa fie la randul 42 cand nu exista manopera"
    assert ws["E42"].value == pytest.approx(50.0), "Subtotal piesă incorect în E42"


def test_grand_total_row_no_labor_1_part(tmp_path):
    """0 labor + 1 billable part: grand total at F44.

    parts_start_row=42, total_parts_row=43, grand_total_row=44.
    """
    part = {"part_id": 10, "part_name": "Filtru ulei", "units": 1.0, "price_per_unit": 200.0}
    ws, _ = _run(
        _base_receipt(
            billable_parts=[part],
            total_parts_cost=200.0,
        ),
        tmp_path,
    )
    assert ws["F44"].value == pytest.approx(200.0)


def test_billable_part_subtotal_with_decimal_units(tmp_path):
    """Decimal quantity (e.g. 1.5 L oil): subtotal = units × price."""
    part = {"part_id": 10, "part_name": "Ulei motor", "units": 1.5, "price_per_unit": 100.0}
    ws, _ = _run(
        _base_receipt(
            labor=[100],
            total_labor_cost=0.0,
            billable_parts=[part],
            total_parts_cost=150.0,
        ),
        tmp_path,
    )
    assert ws["E42"].value == pytest.approx(150.0)


def test_billable_parts_tva_calculated_per_part(tmp_path):
    """TVA per part = (part_total × TVA%) / (100 + TVA%)."""
    part = {"part_id": 10, "part_name": "Filtru ulei", "units": 1.0, "price_per_unit": 121.0}
    expected_tva = (121.0 * TVA) / (100 + TVA)
    ws, _ = _run(
        _base_receipt(
            labor=[100],
            total_labor_cost=0.0,
            billable_parts=[part],
            total_parts_cost=121.0,
        ),
        tmp_path,
    )
    assert ws["F42"].value == pytest.approx(expected_tva, rel=1e-6)


# ---------------------------------------------------------------------------
# Grand total row position  ← critical: this is where bug #2 manifested
# ---------------------------------------------------------------------------


def test_grand_total_row_no_labor_no_parts(tmp_path):
    """Empty receipt (0 labor, 0 parts): grand total written to F44."""
    ws, _ = _run(_base_receipt(), tmp_path)
    assert ws["F44"].value == pytest.approx(0.0)


def test_grand_total_row_1_labor_no_parts(tmp_path):
    """1 labor, 0 billable parts: grand total at F44 (base row, no inserts)."""
    ws, _ = _run(_base_receipt(labor=[100], total_labor_cost=300.0), tmp_path)
    assert ws["F44"].value == pytest.approx(300.0)


def test_grand_total_row_2_labor_no_parts(tmp_path):
    """2 labor items (1 insert): grand total shifts to F45."""
    ws, _ = _run(_base_receipt(labor=[100, 101], total_labor_cost=600.0), tmp_path)
    assert ws["F45"].value == pytest.approx(600.0)


def test_grand_total_row_3_labor_no_parts(tmp_path):
    """3 labor items (2 inserts): grand total shifts to F46."""
    ws, _ = _run(_base_receipt(labor=[100, 101, 102], total_labor_cost=900.0), tmp_path)
    assert ws["F46"].value == pytest.approx(900.0)


def test_grand_total_row_1_labor_1_part(tmp_path):
    """1 labor + 1 billable part: grand total at F44."""
    part = {"part_id": 10, "part_name": "Filtru ulei", "units": 1.0, "price_per_unit": 100.0}
    ws, _ = _run(
        _base_receipt(
            labor=[100],
            total_labor_cost=300.0,
            billable_parts=[part],
            total_parts_cost=100.0,
        ),
        tmp_path,
    )
    assert ws["F44"].value == pytest.approx(400.0)


def test_grand_total_row_2_labor_2_parts(tmp_path):
    """2 labor (1 insert) + 2 billable parts (1 insert): grand total at F46."""
    parts = [
        {"part_id": 10, "part_name": "Filtru ulei", "units": 1.0, "price_per_unit": 50.0},
        {"part_id": 11, "part_name": "Filtru aer", "units": 1.0, "price_per_unit": 30.0},
    ]
    ws, _ = _run(
        _base_receipt(
            labor=[100, 101],
            total_labor_cost=500.0,
            billable_parts=parts,
            total_parts_cost=80.0,
        ),
        tmp_path,
    )
    assert ws["F46"].value == pytest.approx(580.0)


def test_grand_total_value_equals_labor_plus_parts(tmp_path):
    """Grand total = total_labor_cost + total_parts_cost."""
    part = {"part_id": 10, "part_name": "Filtru ulei", "units": 2.0, "price_per_unit": 75.0}
    labor_cost = 400.0
    parts_cost = 150.0
    ws, _ = _run(
        _base_receipt(
            labor=[100],
            total_labor_cost=labor_cost,
            billable_parts=[part],
            total_parts_cost=parts_cost,
        ),
        tmp_path,
    )
    assert ws["F44"].value == pytest.approx(labor_cost + parts_cost)


# ---------------------------------------------------------------------------
# Executant
# ---------------------------------------------------------------------------


def test_executant_at_b50_with_no_extra_rows(tmp_path):
    """No inserts: executant name at B50."""
    ws, _ = _run(_base_receipt(executant_name="Mihai Ionescu"), tmp_path)
    assert ws["B50"].value == "Mihai Ionescu"


def test_executant_shifts_with_extra_labor_and_parts(tmp_path):
    """2 labor (1 extra) + 2 billable parts (1 extra) → executant at B52."""
    parts = [
        {"part_id": 10, "part_name": "Filtru ulei", "units": 1.0, "price_per_unit": 50.0},
        {"part_id": 11, "part_name": "Filtru aer", "units": 1.0, "price_per_unit": 30.0},
    ]
    ws, _ = _run(
        _base_receipt(
            labor=[100, 101],
            total_labor_cost=500.0,
            billable_parts=parts,
            total_parts_cost=80.0,
            executant_name="Radu Constantin",
        ),
        tmp_path,
    )
    assert ws["B52"].value == "Radu Constantin"


# ---------------------------------------------------------------------------
# FULL RECEIPT – toate câmpurile completate
# ---------------------------------------------------------------------------
#
# Configurație:
#   - 2 defecte client          → A14, A15
#   - 2 defecte descoperite     → A20, A21
#   - 2 piese client            → C14, C15
#   - 2 servicii manoperă       → B36, B37  (1 insert extra)
#   - 2 piese facturabile       → B43, B44  (1 insert extra; parts_start_row=43)
#
# Rânduri derivate (cu 2 labor inserts−1=1 și 2 parts inserts−1=1):
#   E38  – total manoperă       (36 + 2)
#   F38  – TVA manoperă
#   E45  – total piese          (43 + 2)
#   F45  – TVA piese
#   F46  – total general        (total_parts_row + 1 = 45 + 1)
#   B52  – executant            (50 + 1 + 1)
# ---------------------------------------------------------------------------


def test_full_receipt_all_fields(tmp_path):
    """
    Deviz complet cu toate câmpurile completate.
    Verifică fiecare celulă relevantă din fișierul Excel generat.
    """
    labor_cost = 450.0
    part1_units, part1_price = 2.0, 100.0  # total 200
    part2_units, part2_price = 1.5, 80.0  # total 120
    parts_cost = part1_units * part1_price + part2_units * part2_price  # 320.0
    grand_total = labor_cost + parts_cost  # 770.0

    receipt = _base_receipt(
        # ── Client / mașină ──────────────────────────────────────────────
        client_name="Alexandru Georgescu",
        client_address="Str. Libertatii 42, Cluj-Napoca",
        model="Skoda Octavia",
        plate_number="CJ99ZZZ",
        vin="TMBZZZ1Z3J1234567",
        kilometers="88500",
        date="07.05.2026",
        executant_name="Ion Mecanicescu",
        # ── Estimări ─────────────────────────────────────────────────────
        estimate_cost=800.0,
        estimated_final_date="12.05.2026",
        # ── Defecte client (2) ───────────────────────────────────────────
        defects=[1, 2],
        # ── Defecte descoperite (2) ──────────────────────────────────────
        discovered_defects=[3, 4],
        # ── Piese furnizate de client (2) ────────────────────────────────
        parts=[10, 11],
        # ── Manoperă (2 servicii) ────────────────────────────────────────
        labor=[100, 101],
        total_labor_cost=labor_cost,
        # ── Piese facturabile (2) ────────────────────────────────────────
        billable_parts=[
            {
                "part_id": 10,
                "part_name": "Filtru ulei",
                "units": part1_units,
                "price_per_unit": part1_price,
            },
            {
                "part_id": 11,
                "part_name": "Ulei motor 5W40",
                "units": part2_units,
                "price_per_unit": part2_price,
            },
        ],
        total_parts_cost=parts_cost,
        grand_total=grand_total,
    )

    ws, warnings = _run(receipt, tmp_path, receipt_number=12)

    # ── Nu trebuie să existe warnings (datele sunt în limite) ────────────
    assert warnings == [], f"Nu ar trebui warnings, dar s-au primit: {warnings}"

    # ── Câmpuri client / mașină ──────────────────────────────────────────
    assert ws["B10"].value == "Alexandru Georgescu", "Nume client incorect în B10"
    assert ws["B12"].value == "Str. Libertatii 42, Cluj-Napoca", "Adresă incorectă în B12"
    assert ws["D10"].value == "Skoda Octavia", "Model mașină incorect în D10"
    assert ws["E11"].value == "CJ99ZZZ", "Număr înmatriculare incorect în E11"
    assert ws["E12"].value == "TMBZZZ1Z3J1234567", "VIN incorect în E12"
    assert ws["F10"].value == 88500, "Kilometraj incorect în F10"

    # ── Număr deviz și dată ──────────────────────────────────────────────
    b8 = str(ws["B8"].value)
    assert "12" in b8, "Numărul devizului (12) lipsește din B8"
    assert "07.05.2026" in b8, "Data lipsește din B8"

    # ── Estimări (A29, A31) ──────────────────────────────────────────────
    assert "800" in str(ws["A29"].value), "Valoarea estimată lipsește din A29"
    assert "12.05.2026" in str(ws["A31"].value), "Data estimată lipsește din A31"

    # ── Defecte client → A14, A15 ────────────────────────────────────────
    assert ws["A14"].value == "Placute frana uzate", "Defect 1 incorect în A14"
    assert ws["A15"].value == "Scurgere ulei motor", "Defect 2 incorect în A15"

    # ── Defecte descoperite → A20, A21 ───────────────────────────────────
    assert ws["A20"].value == "Presiune scazuta anvelope", "Defect descoperit 1 incorect în A20"
    assert ws["A21"].value == "Nivel lichid racire scazut", "Defect descoperit 2 incorect în A21"

    # ── Piese client → C14, C15 ──────────────────────────────────────────
    assert ws["C14"].value == "Filtru ulei", "Piesă client 1 incorectă în C14"
    assert ws["C15"].value == "Filtru aer", "Piesă client 2 incorectă în C15"

    # ── Manoperă → B36, B37; total → E38; TVA → F38 ─────────────────────
    # Header 'Operatie' trebuie sa fie la B34 (2 randuri deasupra B36)
    assert ws["B34"].value == "Operatie", "Header 'Operatie' incorect/absent in B34"
    assert ws["B36"].value == "Schimb ulei", "Manoperă 1 incorectă în B36"
    assert ws["B37"].value == "Inlocuire placute frana", "Manoperă 2 incorectă în B37"
    assert ws["E38"].value == pytest.approx(labor_cost), "Total manoperă incorect în E38"
    expected_labor_tva = (labor_cost * TVA) / (100 + TVA)
    assert ws["F38"].value == pytest.approx(
        expected_labor_tva, rel=1e-6
    ), "TVA manoperă incorect în F38"

    # ── Piese facturabile → B43, B44 ─────────────────────────────────────
    # parts_start_row = 42 + max(0, 2-1) = 43
    # Header 'Piese auto-materiale' trebuie sa fie la B41 (2 randuri deasupra B43)
    assert (
        ws["B41"].value == "Piese auto-materiale"
    ), "Header 'Piese auto-materiale' incorect/absent in B41"
    assert ws["B43"].value == "Filtru ulei", "Piesă facturabilă 1 incorectă în B43"
    assert ws["C43"].value == pytest.approx(part1_units), "Cantitate piesă 1 incorectă în C43"
    assert ws["D43"].value == pytest.approx(part1_price), "Preț/unitate piesă 1 incorect în D43"
    assert ws["E43"].value == pytest.approx(
        part1_units * part1_price
    ), "Subtotal piesă 1 incorect în E43"

    assert ws["B44"].value == "Ulei motor 5W40", "Piesă facturabilă 2 incorectă în B44"
    assert ws["C44"].value == pytest.approx(part2_units), "Cantitate piesă 2 incorectă în C44"
    assert ws["D44"].value == pytest.approx(part2_price), "Preț/unitate piesă 2 incorect în D44"
    assert ws["E44"].value == pytest.approx(
        part2_units * part2_price
    ), "Subtotal piesă 2 incorect în E44"

    # ── Total piese → E45; TVA piese → F45 ───────────────────────────────
    # total_parts_row = parts_start_row + len(billable_parts) = 43 + 2 = 45
    assert ws["E45"].value == pytest.approx(parts_cost), "Total piese incorect în E45"
    expected_parts_tva = (parts_cost * TVA) / (100 + TVA)
    assert ws["F45"].value == pytest.approx(
        expected_parts_tva, rel=1e-6
    ), "TVA piese incorect în F45"

    # ── Total general → F46 ──────────────────────────────────────────────
    # grand_total_row = total_parts_row + 1 = 45 + 1 = 46
    assert ws["F46"].value == pytest.approx(grand_total), "Total general incorect în F46"

    # ── Executant → B52 ──────────────────────────────────────────────────
    # executant_row = 50 + (2-1) + (2-1) = 52
    assert ws["B52"].value == "Ion Mecanicescu", "Executant incorect în B52"
