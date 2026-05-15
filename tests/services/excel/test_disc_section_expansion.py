"""Tests for dynamic row expansion in the discovered-defects section (A19-A24 baseline).

Template layout of the discovered-defects box:
  Row 19  A = "Defecte constatate in timpul reparatiei"   C-F top-border (right panel top)
  Row 20  A = <disc slot 1>                               C = "Accept lucrarile suplimentar constatate."
  Row 21  A = <disc slot 2>                               C = "Semnatura client:"
  Row 22  A = <disc slot 3>                               C = empty (right panel body)
  Row 23  A = <disc slot 4>                               C = empty
  Row 24  A = <disc slot 5>  ← bottom border A-F          C = bottom border
  Row 25  "Masina mai prezinta ..." (next section)

When n_discovered > 5, rows are inserted at row 25 + row_offset (AFTER the section
bottom border), pushing "Masina mai prezinta" and everything below down.

Key invariants:
  • "Accept lucrarile suplimentar constatate."  stays at C(20 + row_offset)
    (it is ABOVE the insertion point → not shifted by disc expansion)
  • "Semnatura client:" (discovered)            stays at C(21 + row_offset)
  • Bottom border moves from row 24+row_offset  → row 24+row_offset+disc_extra
  • "Masina mai prezinta" moves from row 25     → row 25 + total_offset
  • MANOPERA                                    → row 32 + total_offset
  • Labor base row                              → row 36 + total_offset

Formulas:
  disc_extra_rows = max(0, n_discovered - 5)
  total_offset    = row_offset + disc_extra_rows
"""

from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from src.services.excel_export import generate_receipt_excel

# ---------------------------------------------------------------------------
# Mock data
# ---------------------------------------------------------------------------

TVA = 21.0

MOCK_DEFECTS = {i: {"id": i, "defect_name": f"Disc {i}"} for i in range(1, 9)}
MOCK_CLIENT_PARTS = {}
MOCK_LABOR = {100: {"id": 100, "service_name": "Schimb ulei"}}


def _run(receipt_data: dict, tmp_path: Path):
    with (
        patch("src.services.excel_export.EXPORTS_DIR", tmp_path),
        patch("src.services.excel_export.get_tva", return_value=TVA),
        patch("src.services.excel_export.get_receipt_number", return_value=1),
        patch("src.services.excel_export.update_receipt_number"),
        patch(
            "src.services.excel_export.get_defect_by_id",
            side_effect=lambda i: MOCK_DEFECTS.get(i),
        ),
        patch(
            "src.services.excel_export.get_part_by_id",
            side_effect=lambda i: MOCK_CLIENT_PARTS.get(i),
        ),
        patch(
            "src.services.excel_export.get_labor_by_id",
            side_effect=lambda i: MOCK_LABOR.get(i),
        ),
    ):
        output_path, warnings = generate_receipt_excel(receipt_data)

    wb = load_workbook(output_path)
    ws = wb.active
    return ws, warnings


def _base(**overrides) -> dict:
    base = {
        "client_name": "Test Client",
        "client_address": "",
        "model": "Test Car",
        "plate_number": "XX00TST",
        "vin": "",
        "kilometers": "0",
        "executant_name": "",
        "date": "08.05.2026",
        "estimate_cost": 0.0,
        "estimated_final_date": "",
        "defects": [],
        "discovered_defects": [],
        "parts": [],
        "labor": [],
        "total_labor_cost": 0.0,
        "billable_parts": [],
        "total_parts_cost": 0.0,
        "grand_total": 0.0,
    }
    base.update(overrides)
    return base


def _has_medium_bottom(cell) -> bool:
    b = cell.border
    return bool(b.bottom and b.bottom.style == "medium")


# ===========================================================================
# Baseline: n_discovered <= 5  →  no expansion, existing behaviour unchanged
# ===========================================================================


class TestDiscoveredBaselineUnchanged:
    """Sanity: five or fewer discovered defects must not trigger expansion."""

    def test_five_discovered_no_extra_rows(self, tmp_path):
        """Five discovered defects fill A20-A24 with no insertion."""
        ws, _ = _run(_base(discovered_defects=list(range(1, 6))), tmp_path)
        for i in range(5):
            assert (
                ws[f"A{20 + i}"].value == f"Disc {i + 1}"
            ), f"A{20 + i}: {ws[f'A{20 + i}'].value!r}"

    def test_five_discovered_no_warning(self, tmp_path):
        ws, warnings = _run(_base(discovered_defects=list(range(1, 6))), tmp_path)
        assert warnings == [], f"Unexpected warnings: {warnings}"

    def test_accept_lucrarile_at_c20_baseline(self, tmp_path):
        """'Accept lucrarile...' remains at C20 when no expansion occurs."""
        ws, _ = _run(_base(), tmp_path)
        assert (
            ws["C20"].value == "Accept lucrarile suplimentar constatate."
        ), f"C20: {ws['C20'].value!r}"

    def test_semnatura_disc_at_c21_baseline(self, tmp_path):
        """'Semnatura client:' (discovered) remains at C21 at baseline."""
        ws, _ = _run(_base(), tmp_path)
        assert ws["C21"].value == "Semnatura client:", f"C21: {ws['C21'].value!r}"

    def test_bottom_border_at_row_24_baseline(self, tmp_path):
        """Bottom border of the disc section stays at row 24 with no expansion."""
        ws, _ = _run(_base(), tmp_path)
        for col in ["A", "B", "C", "D", "E", "F"]:
            assert _has_medium_bottom(
                ws[f"{col}24"]
            ), f"{col}24 should have medium bottom border (baseline)"


# ===========================================================================
# n_discovered = 8 with NO client section expansion  (row_offset = 0)
# disc_extra_rows = 3,  total_offset = 3
# ===========================================================================


class TestDiscExpansionNoClientExpansion:
    """
    8 discovered defects, no client defects/parts.
    row_offset = 0,  disc_extra_rows = 3,  total_offset = 3.

    Expected layout:
      Disc slots         : A20 – A27
      Accept lucrarile   : C20  (unchanged – above insertion point)
      Semnatura client   : C21  (unchanged – above insertion point)
      Bottom border      : row 27  (= 24 + 3)
      Masina mai prezinta: A28  (= 25 + 3)
      MANOPERA           : A35  (= 32 + 3)
      Labor base         : B39  (= 36 + 3, with 1 labor item)
    """

    N = 8
    IDS = list(range(1, 9))
    DISC_EXTRA = 3  # max(0, 8-5)
    OFFSET = 0  # row_offset (no client expansion)
    TOTAL = DISC_EXTRA + OFFSET  # 3

    def test_all_discovered_defects_written(self, tmp_path):
        """All 8 discovered defects written to A20-A27 with no truncation."""
        ws, _ = _run(_base(discovered_defects=self.IDS), tmp_path)
        for i, did in enumerate(self.IDS):
            expected = MOCK_DEFECTS[did]["defect_name"]
            row = 20 + self.OFFSET + i
            assert ws[f"A{row}"].value == expected, (
                f"Disc defect {i + 1}: A{row} expected {expected!r}, "
                f"got {ws[f'A{row}'].value!r}"
            )

    def test_no_overflow_warning_when_section_expands(self, tmp_path):
        """No warning emitted when section expands to fit discovered defects."""
        _, warnings = _run(_base(discovered_defects=self.IDS), tmp_path)
        assert warnings == [], f"Unexpected warnings: {warnings}"

    def test_accept_lucrarile_stays_in_place(self, tmp_path):
        """'Accept lucrarile...' must remain at C20 (above insertion point)."""
        ws, _ = _run(_base(discovered_defects=self.IDS), tmp_path)
        row = 20 + self.OFFSET
        assert (
            ws[f"C{row}"].value == "Accept lucrarile suplimentar constatate."
        ), f"C{row}: {ws[f'C{row}'].value!r}"

    def test_semnatura_disc_stays_in_place(self, tmp_path):
        """'Semnatura client:' (discovered) must remain at C21 (above insertion point)."""
        ws, _ = _run(_base(discovered_defects=self.IDS), tmp_path)
        row = 21 + self.OFFSET
        assert ws[f"C{row}"].value == "Semnatura client:", f"C{row}: {ws[f'C{row}'].value!r}"

    def test_bottom_border_moved_to_correct_row(self, tmp_path):
        """Bottom border must be at row 27 (= 24 + disc_extra_rows)."""
        ws, _ = _run(_base(discovered_defects=self.IDS), tmp_path)
        new_border_row = 24 + self.OFFSET + self.DISC_EXTRA  # 27
        for col in ["A", "B", "C", "D", "E", "F"]:
            assert _has_medium_bottom(
                ws[f"{col}{new_border_row}"]
            ), f"{col}{new_border_row} should have medium bottom border"

    def test_bottom_border_removed_from_original_row(self, tmp_path):
        """Medium bottom border must NOT remain at old row 24 after expansion."""
        ws, _ = _run(_base(discovered_defects=self.IDS), tmp_path)
        for col in ["C", "D", "E", "F"]:
            assert not _has_medium_bottom(
                ws[f"{col}24"]
            ), f"{col}24 still has medium bottom border after expansion"

    def test_masina_mai_prezinta_shifts_down(self, tmp_path):
        """'Masina mai prezinta...' must be at A(25 + total_offset) = A28."""
        ws, _ = _run(_base(discovered_defects=self.IDS), tmp_path)
        row = 25 + self.TOTAL
        assert "Masina mai prezinta" in (
            ws[f"A{row}"].value or ""
        ), f"A{row}: {ws[f'A{row}'].value!r}"

    def test_manopera_shifts_by_total_offset(self, tmp_path):
        """MANOPERA header must be at A(32 + total_offset) = A35."""
        ws, _ = _run(_base(discovered_defects=self.IDS), tmp_path)
        row = 32 + self.TOTAL
        assert ws[f"A{row}"].value == "MANOPERA", f"A{row}: {ws[f'A{row}'].value!r}"

    def test_labor_shifts_by_total_offset(self, tmp_path):
        """With 1 labor item, it must be at B(36 + total_offset) = B39."""
        ws, _ = _run(
            _base(discovered_defects=self.IDS, labor=[100], total_labor_cost=300.0),
            tmp_path,
        )
        row = 36 + self.TOTAL
        assert ws[f"B{row}"].value == "Schimb ulei", f"B{row}: {ws[f'B{row}'].value!r}"


# ===========================================================================
# n_discovered = 8 WITH client section expansion  (row_offset = 3, 8 defects)
# disc_extra_rows = 3,  total_offset = 6
# ===========================================================================


class TestDiscExpansionWithClientExpansion:
    """
    8 discovered defects + 8 client defects (row_offset = 3).
    disc_extra_rows = 3,  total_offset = 6.

    Expected layout:
      Disc section header: A22  (= 19 + row_offset)
      Disc slots         : A23 – A30  (= 20+3 .. 19+3+8)
      Accept lucrarile   : C23  (= 20 + row_offset, NOT shifted by disc)
      Semnatura client   : C24  (= 21 + row_offset, NOT shifted by disc)
      Bottom border      : row 30  (= 24 + 3 + 3)
      Masina mai prezinta: A31  (= 25 + 6)
      MANOPERA           : A38  (= 32 + 6)
      Labor base         : B42  (= 36 + 6, with 1 labor item)
    """

    N_CLIENT = 8
    N_DISC = 8
    CLIENT_IDS = list(range(1, 9))
    DISC_IDS = list(range(1, 9))
    ROW_OFFSET = 3  # from client section (max(5,8,0+2)-5 = 3)
    DISC_EXTRA = 3  # max(0, 8-5)
    TOTAL = ROW_OFFSET + DISC_EXTRA  # 6

    def test_all_discovered_defects_written_at_shifted_rows(self, tmp_path):
        """All 8 disc defects written to A23-A30 (shifted by row_offset=3)."""
        ws, _ = _run(_base(defects=self.CLIENT_IDS, discovered_defects=self.DISC_IDS), tmp_path)
        for i, did in enumerate(self.DISC_IDS):
            expected = MOCK_DEFECTS[did]["defect_name"]
            row = 20 + self.ROW_OFFSET + i  # 23..30
            assert ws[f"A{row}"].value == expected, (
                f"Disc defect {i + 1}: A{row} expected {expected!r}, "
                f"got {ws[f'A{row}'].value!r}"
            )

    def test_no_overflow_warning_combined(self, tmp_path):
        """No warnings when both sections expand."""
        _, warnings = _run(
            _base(defects=self.CLIENT_IDS, discovered_defects=self.DISC_IDS), tmp_path
        )
        assert warnings == [], f"Unexpected warnings: {warnings}"

    def test_accept_lucrarile_shifted_by_client_offset_only(self, tmp_path):
        """'Accept lucrarile...' at C(20 + row_offset) = C23, NOT shifted by disc expansion."""
        ws, _ = _run(_base(defects=self.CLIENT_IDS, discovered_defects=self.DISC_IDS), tmp_path)
        row = 20 + self.ROW_OFFSET  # C23
        assert (
            ws[f"C{row}"].value == "Accept lucrarile suplimentar constatate."
        ), f"C{row}: {ws[f'C{row}'].value!r}"

    def test_semnatura_disc_shifted_by_client_offset_only(self, tmp_path):
        """'Semnatura client:' (disc) at C(21 + row_offset) = C24, NOT shifted by disc."""
        ws, _ = _run(_base(defects=self.CLIENT_IDS, discovered_defects=self.DISC_IDS), tmp_path)
        row = 21 + self.ROW_OFFSET  # C24
        assert ws[f"C{row}"].value == "Semnatura client:", f"C{row}: {ws[f'C{row}'].value!r}"

    def test_bottom_border_with_both_expansions(self, tmp_path):
        """Bottom border at row (24 + row_offset + disc_extra) = row 30."""
        ws, _ = _run(_base(defects=self.CLIENT_IDS, discovered_defects=self.DISC_IDS), tmp_path)
        border_row = 24 + self.ROW_OFFSET + self.DISC_EXTRA  # 30
        for col in ["A", "B", "C", "D", "E", "F"]:
            assert _has_medium_bottom(
                ws[f"{col}{border_row}"]
            ), f"{col}{border_row} should have medium bottom border"

    def test_masina_mai_prezinta_with_both_expansions(self, tmp_path):
        """'Masina mai prezinta...' at A(25 + total_offset) = A31."""
        ws, _ = _run(_base(defects=self.CLIENT_IDS, discovered_defects=self.DISC_IDS), tmp_path)
        row = 25 + self.TOTAL  # 31
        assert "Masina mai prezinta" in (
            ws[f"A{row}"].value or ""
        ), f"A{row}: {ws[f'A{row}'].value!r}"

    def test_manopera_with_both_expansions(self, tmp_path):
        """MANOPERA at A(32 + total_offset) = A38."""
        ws, _ = _run(_base(defects=self.CLIENT_IDS, discovered_defects=self.DISC_IDS), tmp_path)
        row = 32 + self.TOTAL  # 38
        assert ws[f"A{row}"].value == "MANOPERA", f"A{row}: {ws[f'A{row}'].value!r}"

    def test_labor_with_both_expansions(self, tmp_path):
        """With 1 labor item, it must be at B(36 + total_offset) = B42."""
        ws, _ = _run(
            _base(
                defects=self.CLIENT_IDS,
                discovered_defects=self.DISC_IDS,
                labor=[100],
                total_labor_cost=300.0,
            ),
            tmp_path,
        )
        row = 36 + self.TOTAL  # 42
        assert ws[f"B{row}"].value == "Schimb ulei", f"B{row}: {ws[f'B{row}'].value!r}"


# ===========================================================================
# Discovered section border behaviour at various sizes
# ===========================================================================


class TestDiscSectionBorderLogic:
    """Bottom border movement at the exact expansion thresholds."""

    def test_no_border_move_with_exactly_five(self, tmp_path):
        """n=5: no expansion → bottom border stays at row 24."""
        ws, _ = _run(_base(discovered_defects=list(range(1, 6))), tmp_path)
        for col in ["C", "D", "E", "F"]:
            assert _has_medium_bottom(ws[f"{col}24"]), f"{col}24 should keep medium bottom with n=5"

    def test_border_moves_with_six_discovered(self, tmp_path):
        """n=6: disc_extra=1 → border moves from row 24 to row 25."""
        ws, _ = _run(_base(discovered_defects=list(range(1, 7))), tmp_path)
        for col in ["C", "D", "E", "F"]:
            assert not _has_medium_bottom(
                ws[f"{col}24"]
            ), f"{col}24 should NOT have medium bottom with n=6"
            assert _has_medium_bottom(ws[f"{col}25"]), f"{col}25 should have medium bottom with n=6"

    def test_border_moves_with_eight_discovered(self, tmp_path):
        """n=8: disc_extra=3 → border moves from row 24 to row 27."""
        ws, _ = _run(_base(discovered_defects=list(range(1, 9))), tmp_path)
        for col in ["C", "D", "E", "F"]:
            assert not _has_medium_bottom(
                ws[f"{col}24"]
            ), f"{col}24 should NOT have medium bottom with n=8"
            assert _has_medium_bottom(ws[f"{col}27"]), f"{col}27 should have medium bottom with n=8"

    def test_a_column_bottom_border_moves_with_eight(self, tmp_path):
        """A column bottom border also moves to row 27 with n=8."""
        ws, _ = _run(_base(discovered_defects=list(range(1, 9))), tmp_path)
        assert _has_medium_bottom(ws["A27"]), "A27 should have medium bottom with n=8"
        assert not _has_medium_bottom(ws["A24"]), "A24 should NOT have medium bottom with n=8"
