"""Tests for layout invariants of the generated Excel receipt.

Covers scenarios not exercised by test_export.py or test_section_expansion.py:

1. TestGrandTotalRowWithExpansion
   Grand total row formula must include row_offset.
   These tests DOCUMENT the expected behaviour and expose the bug where
   `grand_total_row = 43 + N` / `44` do not add row_offset.

2. TestDiscoveredDefectsSectionLabels
   Static template labels in the "Defecte constatate" section shift by
   row_offset after section expansion.

3. TestEstimateFields
   Estimate cost (A29) and estimated final date (A31) are written with the
   correct values, both at baseline and after expansion.

4. TestStaticLabelsAfterExpansion
   "Masina mai prezinta" section, MANOPERA column headers, and the
   "Total general de plata" label all land on the correct rows after expansion.

Row-offset formula reminder:
    row_offset = max(5, n_defects, n_parts + 2) - 5
    grand_total_row (no billable parts, N labor):  44 + row_offset + (N-1)  = 43 + N + row_offset
    grand_total_row (no labor, no billable parts): 44 + row_offset
"""

import pytest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from src.services.excel_export import generate_receipt_excel

# ---------------------------------------------------------------------------
# Shared constants
# ---------------------------------------------------------------------------

TVA = 21.0

MOCK_DEFECTS = {i: {"id": i, "defect_name": f"Defect {i}"} for i in range(1, 9)}
MOCK_PARTS   = {i: {"id": i, "part_name": f"Part {i}"}     for i in range(10, 16)}
MOCK_LABOR   = {
    100: {"id": 100, "service_name": "Schimb ulei"},
    101: {"id": 101, "service_name": "Schimb filtre"},
    102: {"id": 102, "service_name": "Schimb bujii"},
}

# 8 defects  →  section_rows=8, extra_rows=3, row_offset=3
MANY_DEFECTS = list(range(1, 9))


def _run(receipt_data: dict, tmp_path: Path):
    """Generate with all DB mocks and return (worksheet, warnings)."""
    with (
        patch("src.services.excel_export.EXPORTS_DIR", tmp_path),
        patch("src.services.excel_export.get_tva", return_value=TVA),
        patch("src.services.excel_export.get_receipt_number", return_value=1),
        patch("src.services.excel_export.update_receipt_number"),
        patch(
            "src.services.excel_export.get_defect_by_id",
            side_effect=lambda i: MOCK_DEFECTS.get(i),
        ),
        patch(
            "src.services.excel_export.get_part_by_id",
            side_effect=lambda i: MOCK_PARTS.get(i),
        ),
        patch(
            "src.services.excel_export.get_labor_by_id",
            side_effect=lambda i: MOCK_LABOR.get(i),
        ),
    ):
        output_path, warnings = generate_receipt_excel(receipt_data)

    wb = load_workbook(output_path)
    ws = wb.active
    return ws, warnings


def _base(**overrides) -> dict:
    base = {
        "client_name": "Test Client",
        "client_address": "Str. Testului 1",
        "model": "Dacia Logan",
        "plate_number": "B01TST",
        "vin": "VIN123",
        "kilometers": "50000",
        "executant_name": "Ion Marius",
        "date": "08.05.2026",
        "estimate_cost": 350.0,
        "estimated_final_date": "15.05.2026",
        "defects": [],
        "discovered_defects": [],
        "parts": [],
        "labor": [],
        "total_labor_cost": 0.0,
        "billable_parts": [],
        "total_parts_cost": 0.0,
        "grand_total": 0.0,
    }
    base.update(overrides)
    return base


# ===========================================================================
# Grand total row with section expansion (row_offset > 0)
# ===========================================================================

class TestGrandTotalRowWithExpansion:
    """
    When the defects/parts section expands (row_offset > 0) and there are no
    billable parts, the grand_total_row formula must include row_offset.

    Using 8 defects → row_offset = 3.
    Expected grand_total rows:
      no labor, no billable:  44 + 3 = 47
      1 labor,  no billable:  43 + 1 + 3 = 47
      2 labor,  no billable:  43 + 2 + 3 = 48
      3 labor,  no billable:  43 + 3 + 3 = 49
    """

    OFFSET = 3

    def test_grand_total_row_expansion_no_labor(self, tmp_path):
        """No labor, no billable parts: grand total at F(44 + row_offset) = F47."""
        expected_row = 44 + self.OFFSET   # 47
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        val = ws[f"F{expected_row}"].value
        assert val == 0.0 or val == 0, (
            f"F{expected_row} should be grand_total=0, got {val!r}"
        )
        # Make sure it was NOT written to the old un-offset row
        wrong_row = 44
        wrong_val = ws[f"F{wrong_row}"].value
        assert wrong_val != 0.0 or str(wrong_val) in ("None", "0.21", "TVA-like"), (
            f"Grand total should NOT be at F{wrong_row} when row_offset={self.OFFSET}"
        )

    def test_grand_total_value_matches_with_expansion_no_labor(self, tmp_path):
        """Grand total VALUE at the correct row equals labor + parts (= 0 here)."""
        expected_row = 44 + self.OFFSET
        ws, _ = _run(_base(defects=MANY_DEFECTS, grand_total=0.0), tmp_path)
        val = ws[f"F{expected_row}"].value
        assert val is not None, f"F{expected_row} is None (not written)"
        assert float(val) == 0.0, f"F{expected_row} expected 0.0, got {val!r}"

    def test_grand_total_row_expansion_1_labor_no_billable(self, tmp_path):
        """1 labor item, no billable parts: grand total at F(43 + 1 + row_offset) = F47."""
        expected_row = 43 + 1 + self.OFFSET   # 47
        ws, _ = _run(
            _base(defects=MANY_DEFECTS, labor=[100], total_labor_cost=200.0, grand_total=200.0),
            tmp_path,
        )
        val = ws[f"F{expected_row}"].value
        assert val is not None, f"F{expected_row} is None (not written)"
        assert float(val) == 200.0, (
            f"F{expected_row} expected grand_total=200.0, got {val!r}"
        )

    def test_grand_total_row_expansion_2_labor_no_billable(self, tmp_path):
        """2 labor items, no billable parts: grand total at F(43 + 2 + row_offset) = F48."""
        expected_row = 43 + 2 + self.OFFSET   # 48
        ws, _ = _run(
            _base(defects=MANY_DEFECTS, labor=[100, 101], total_labor_cost=400.0, grand_total=400.0),
            tmp_path,
        )
        val = ws[f"F{expected_row}"].value
        assert val is not None, f"F{expected_row} is None (not written)"
        assert float(val) == 400.0, (
            f"F{expected_row} expected grand_total=400.0, got {val!r}"
        )

    def test_grand_total_row_expansion_3_labor_no_billable(self, tmp_path):
        """3 labor items, no billable parts: grand total at F(43 + 3 + row_offset) = F49."""
        expected_row = 43 + 3 + self.OFFSET   # 49
        ws, _ = _run(
            _base(defects=MANY_DEFECTS, labor=[100, 101, 102], total_labor_cost=600.0,
                  grand_total=600.0),
            tmp_path,
        )
        val = ws[f"F{expected_row}"].value
        assert val is not None, f"F{expected_row} is None (not written)"
        assert float(val) == 600.0, (
            f"F{expected_row} expected grand_total=600.0, got {val!r}"
        )

    def test_total_general_label_shifts_no_labor(self, tmp_path):
        """'Total general de plata' label must be at D(44 + row_offset) = D47."""
        expected_row = 44 + self.OFFSET   # 47
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        val = ws[f"D{expected_row}"].value
        assert val is not None and "total general" in str(val).lower(), (
            f"D{expected_row} expected 'Total general de plata', got {val!r}"
        )

    def test_total_general_label_shifts_with_labor(self, tmp_path):
        """With 2 labor items, 'Total general' label at D(43 + 2 + row_offset) = D48."""
        expected_row = 43 + 2 + self.OFFSET   # 48
        ws, _ = _run(
            _base(defects=MANY_DEFECTS, labor=[100, 101], total_labor_cost=400.0),
            tmp_path,
        )
        val = ws[f"D{expected_row}"].value
        assert val is not None and "total general" in str(val).lower(), (
            f"D{expected_row} expected 'Total general de plata', got {val!r}"
        )


# ===========================================================================
# Discovered-defects section labels — baseline and after expansion
# ===========================================================================

class TestDiscoveredDefectsSectionLabels:
    """
    Template static labels in the discovered-defects block must remain intact
    after generation — both at baseline (row_offset=0) and with expansion.
    """

    def test_defecte_constatate_header_baseline(self, tmp_path):
        """A19 = 'Defecte constatate in timpul reparatiei' with no expansion."""
        ws, _ = _run(_base(), tmp_path)
        assert ws["A19"].value == "Defecte constatate in timpul reparatiei", (
            f"A19: {ws['A19'].value!r}"
        )

    def test_defecte_constatate_header_shifts_with_expansion(self, tmp_path):
        """After row_offset=3, the header moves to A(19 + 3) = A22."""
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        assert ws["A22"].value == "Defecte constatate in timpul reparatiei", (
            f"A22: {ws['A22'].value!r}"
        )

    def test_accept_lucrarile_baseline(self, tmp_path):
        """C20 = 'Accept lucrarile suplimentar constatate.' with no expansion."""
        ws, _ = _run(_base(), tmp_path)
        assert ws["C20"].value == "Accept lucrarile suplimentar constatate.", (
            f"C20: {ws['C20'].value!r}"
        )

    def test_accept_lucrarile_shifts_with_expansion(self, tmp_path):
        """After row_offset=3, 'Accept lucrarile...' moves to C(20 + 3) = C23."""
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        assert ws["C23"].value == "Accept lucrarile suplimentar constatate.", (
            f"C23: {ws['C23'].value!r}"
        )

    def test_semnatura_discovered_baseline(self, tmp_path):
        """C21 = 'Semnatura client:' (discovered-defects signature) with no expansion."""
        ws, _ = _run(_base(), tmp_path)
        assert ws["C21"].value == "Semnatura client:", (
            f"C21: {ws['C21'].value!r}"
        )

    def test_semnatura_discovered_shifts_with_expansion(self, tmp_path):
        """After row_offset=3, the discovered-defects signature moves to C(21 + 3) = C24."""
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        assert ws["C24"].value == "Semnatura client:", (
            f"C24: {ws['C24'].value!r}"
        )

    def test_discovered_defect_slot_1_baseline(self, tmp_path):
        """Discovered defect 1 written to A20 (first slot below section header)."""
        ws, _ = _run(_base(discovered_defects=[1, 2]), tmp_path)
        assert ws["A20"].value == "Defect 1", f"A20: {ws['A20'].value!r}"

    def test_discovered_defect_slot_1_with_expansion(self, tmp_path):
        """After row_offset=3, discovered defect 1 written to A(20 + 3) = A23."""
        ws, _ = _run(_base(defects=MANY_DEFECTS, discovered_defects=[1, 2]), tmp_path)
        assert ws["A23"].value == "Defect 1", f"A23: {ws['A23'].value!r}"
        assert ws["A24"].value == "Defect 2", f"A24: {ws['A24'].value!r}"


# ===========================================================================
# Estimate / time fields — content and position
# ===========================================================================

class TestEstimateFields:
    """
    estimate_cost → embedded in A29 string.
    estimated_final_date → embedded in A31 string.
    Both cells shift by row_offset after section expansion.
    """

    ESTIMATE_COST = 350.0
    FINAL_DATE    = "15.05.2026"

    def test_estimate_cost_in_a29_baseline(self, tmp_path):
        """A29 contains the estimate_cost value at baseline (no expansion)."""
        ws, _ = _run(_base(estimate_cost=self.ESTIMATE_COST), tmp_path)
        val = ws["A29"].value or ""
        assert "350.0" in str(val) or "350" in str(val), (
            f"A29 should contain estimate_cost=350; got: {val!r}"
        )

    def test_estimated_final_date_in_a31_baseline(self, tmp_path):
        """A31 contains the estimated_final_date string at baseline."""
        ws, _ = _run(_base(estimated_final_date=self.FINAL_DATE), tmp_path)
        val = ws["A31"].value or ""
        assert self.FINAL_DATE in str(val), (
            f"A31 should contain '{self.FINAL_DATE}'; got: {val!r}"
        )

    def test_estimate_cost_shifts_with_expansion(self, tmp_path):
        """After row_offset=3, estimate cost shifts to A(29 + 3) = A32."""
        ws, _ = _run(_base(defects=MANY_DEFECTS, estimate_cost=self.ESTIMATE_COST), tmp_path)
        val = ws["A32"].value or ""
        assert "350" in str(val), (
            f"A32 should contain estimate_cost=350 after expansion; got: {val!r}"
        )

    def test_estimated_final_date_shifts_with_expansion(self, tmp_path):
        """After row_offset=3, final date shifts to A(31 + 3) = A34."""
        ws, _ = _run(
            _base(defects=MANY_DEFECTS, estimated_final_date=self.FINAL_DATE), tmp_path
        )
        val = ws["A34"].value or ""
        assert self.FINAL_DATE in str(val), (
            f"A34 should contain '{self.FINAL_DATE}' after expansion; got: {val!r}"
        )


# ===========================================================================
# Static section labels — "Masina mai prezinta" and MANOPERA column headers
# ===========================================================================

class TestStaticLabelsAfterExpansion:
    """
    Template labels that are never overwritten by the code should survive
    generation unchanged and shift correctly when rows are inserted.
    """

    def test_masina_mai_prezinta_baseline(self, tmp_path):
        """A25 = 'Masina mai prezinta urmatoarele deficiente:' at baseline."""
        ws, _ = _run(_base(), tmp_path)
        assert "Masina mai prezinta" in (ws["A25"].value or ""), (
            f"A25: {ws['A25'].value!r}"
        )

    def test_masina_mai_prezinta_shifts_with_expansion(self, tmp_path):
        """After row_offset=3, 'Masina mai prezinta...' moves to A(25 + 3) = A28."""
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        assert "Masina mai prezinta" in (ws["A28"].value or ""), (
            f"A28: {ws['A28'].value!r}"
        )

    def test_manopera_column_headers_baseline(self, tmp_path):
        """MANOPERA table headers at A34–F34 are intact at baseline."""
        ws, _ = _run(_base(), tmp_path)
        assert ws["A34"].value == "Nr.",         f"A34: {ws['A34'].value!r}"
        assert ws["B34"].value == "Operatie",    f"B34: {ws['B34'].value!r}"
        assert ws["C34"].value == "Timp",        f"C34: {ws['C34'].value!r}"
        assert ws["D34"].value == "Pret unitar", f"D34: {ws['D34'].value!r}"
        assert ws["E34"].value == "Valoare",     f"E34: {ws['E34'].value!r}"
        assert ws["F34"].value == "TVA",         f"F34: {ws['F34'].value!r}"

    def test_manopera_column_headers_shift_with_expansion(self, tmp_path):
        """After row_offset=3, MANOPERA table headers move to row 37."""
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        r = 34 + 3   # 37
        assert ws[f"A{r}"].value == "Nr.",         f"A{r}: {ws[f'A{r}'].value!r}"
        assert ws[f"B{r}"].value == "Operatie",    f"B{r}: {ws[f'B{r}'].value!r}"
        assert ws[f"D{r}"].value == "Pret unitar", f"D{r}: {ws[f'D{r}'].value!r}"
        assert ws[f"E{r}"].value == "Valoare",     f"E{r}: {ws[f'E{r}'].value!r}"
        assert ws[f"F{r}"].value == "TVA",         f"F{r}: {ws[f'F{r}'].value!r}"

    def test_piese_column_headers_baseline(self, tmp_path):
        """PIESE table headers at A40–F40 are intact at baseline."""
        ws, _ = _run(_base(), tmp_path)
        assert ws["A40"].value == "Nr.",                  f"A40: {ws['A40'].value!r}"
        assert ws["B40"].value == "Piese auto-materiale", f"B40: {ws['B40'].value!r}"
        assert ws["C40"].value == "Buc.",                 f"C40: {ws['C40'].value!r}"
        assert ws["D40"].value == "Pret unitar",          f"D40: {ws['D40'].value!r}"

    def test_piese_column_headers_shift_with_expansion(self, tmp_path):
        """After row_offset=3, PIESE table headers move to row 43."""
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        r = 40 + 3   # 43
        assert ws[f"A{r}"].value == "Nr.",                  f"A{r}: {ws[f'A{r}'].value!r}"
        assert ws[f"B{r}"].value == "Piese auto-materiale", f"B{r}: {ws[f'B{r}'].value!r}"
        assert ws[f"C{r}"].value == "Buc.",                 f"C{r}: {ws[f'C{r}'].value!r}"

    def test_total_general_label_baseline(self, tmp_path):
        """D44 = 'Total general de plata' at baseline (no expansion, no labor)."""
        ws, _ = _run(_base(), tmp_path)
        val = ws["D44"].value or ""
        assert "total general" in val.lower(), f"D44: {val!r}"

    def test_total_general_label_shifts_no_labor(self, tmp_path):
        """After row_offset=3, 'Total general' label at D(44 + 3) = D47."""
        ws, _ = _run(_base(defects=MANY_DEFECTS), tmp_path)
        val = ws["D47"].value or ""
        assert "total general" in val.lower(), f"D47: {val!r}"

    def test_total_general_label_shifts_with_labor(self, tmp_path):
        """After row_offset=3 and 2 labor items, label at D(43 + 2 + 3) = D48."""
        ws, _ = _run(
            _base(defects=MANY_DEFECTS, labor=[100, 101], total_labor_cost=300.0),
            tmp_path,
        )
        val = ws["D48"].value or ""
        assert "total general" in val.lower(), f"D48: {val!r}"
