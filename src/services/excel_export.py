"""Excel export functionality for receipts."""

import shutil
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Border, Side
from openpyxl.utils.units import cm_to_EMU

from src.database.models.defects import get_defect_by_id
from src.database.models.labor import get_labor_by_id
from src.database.models.parts import get_part_by_id
from src.database.models.settings import get_receipt_number, get_tva, update_receipt_number
from src.paths import get_app_dir, get_bundle_dir

# Templates are bundled read-only assets; exports live next to the exe/project
TEMPLATES_DIR = get_bundle_dir() / "templates"
EXPORTS_DIR = get_app_dir() / "exports" / "receipts"

# Template file name
TEMPLATE_FILE = "Template-deviz.xlsx"
LOGO_FILE = TEMPLATES_DIR / "images" / "Neloaica_logo.png"


def ensure_exports_dir():
    """Ensure the exports directory exists."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def generate_receipt_excel(receipt_data: dict) -> str:
    """
    Generate a receipt Excel file from the template.

    Args:
        receipt_data: Dictionary containing all receipt information

    Returns:
        Path to the generated Excel file

    Raises:
        FileNotFoundError: If template file is not found
        Exception: If there's an error generating the file
    """
    # Ensure exports directory exists
    ensure_exports_dir()

    template_path = TEMPLATES_DIR / TEMPLATE_FILE
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    # Generate filename: deviz_<receipt_number>_<plate_number>_<year>
    receipt_number = get_receipt_number()
    plate_number = receipt_data.get("plate_number", "Unknown").replace(" ", "_")
    year = datetime.now().strftime("%Y")
    output_filename = f"deviz_{receipt_number}_{plate_number}_{year}.xlsx"
    output_path = EXPORTS_DIR / output_filename

    # Copy template to output location
    shutil.copy2(template_path, output_path)

    # Open the copied file and fill in the data
    workbook = load_workbook(output_path)
    sheet = workbook.active  # Get Sheet 1

    # Insert logo into D5:E9 area (temp copy avoids OneDrive file locks)
    tmp_logo = None
    if LOGO_FILE.exists():
        tmp_logo = Path(tempfile.gettempdir()) / "neloaica_logo_tmp.png"
        shutil.copy2(LOGO_FILE, tmp_logo)
        logo = XlImage(str(tmp_logo))
        # Target box D5:E9 is roughly 4.5 cm wide × 2.8 cm tall.
        # Logo native aspect ratio is ~2.19:1 (1287×587).
        # Scale to 4.0 cm wide → height = 4.0 / 2.19 ≈ 1.83 cm, centered with margin.
        logo_w_cm = 6.0
        logo_h_cm = logo_w_cm * logo.height / logo.width
        logo.width = cm_to_EMU(logo_w_cm) / 9525
        logo.height = cm_to_EMU(logo_h_cm) / 9525
        sheet.add_image(logo, "D5")

    # B8: Receipt number and date
    receipt_date = receipt_data.get("date", "")
    sheet["B8"] = (
        f"                                     Nr. {receipt_number}   Din data {receipt_date}"
    )

    # Increment receipt number for the next receipt
    update_receipt_number(receipt_number + 1)

    # A1: Estimated repair value
    estimated_cost = receipt_data.get("estimate_cost", 0.0)
    sheet["A29"] = (
        f"Valoarea estimativa reparatie:    {estimated_cost} LEI"
        "                 Intocmit:  CORAS DIANA                        Accept client:"
    )

    # A31: Estimated repair time
    estimated_final_date = receipt_data.get("estimated_final_date", "")
    sheet["A31"] = (
        f"Timp estimativ reparatie:      {estimated_final_date}"
        "                                Semnatura:                              Semnatura:"
    )

    # Fill in client data
    # B10: Client name
    sheet["B10"] = receipt_data.get("client_name", "")

    # D10: Car model
    sheet["D10"] = receipt_data.get("model", "")

    # F10: Car kilometers
    kilometers = receipt_data.get("kilometers", "")
    if kilometers:
        try:
            sheet["F10"] = int(kilometers)
        except ValueError:
            sheet["F10"] = kilometers

    # E11: Plate number
    sheet["E11"] = receipt_data.get("plate_number", "")

    # E12: VIN
    sheet["E12"] = receipt_data.get("vin", "")

    # B12: Client address
    sheet["B12"] = receipt_data.get("client_address", "")

    # ── Expandable defects / client-parts section (rows 13-18 baseline) ──────
    # Template minimum: 5 data rows (rows 14-18).
    #   Left  (defects): N_defects slots
    #   Right (parts + signature): N_parts slots + 2 fixed signature rows
    # If either side overflows, insert extra rows at row 17 (just before the
    # signature rows). openpyxl shifts everything below that point automatically.
    defect_ids = receipt_data.get("defects", [])
    part_ids = receipt_data.get("parts", [])
    warning_messages = []

    # Reusable border sides for the two expandable sections
    _medium = Side(style="medium")
    _none = Side(style=None)

    n_defects = len(defect_ids)
    n_parts = len(part_ids)
    section_rows = max(5, n_defects, n_parts + 2)
    extra_rows = section_rows - 5

    if extra_rows > 0:
        sheet.insert_rows(17, amount=extra_rows)
        # Copy formatting from row 14 (generic data row) to each inserted row
        for new_row in range(17, 17 + extra_rows):
            for col in range(1, sheet.max_column + 1):
                source = sheet.cell(row=14, column=col)
                target = sheet.cell(row=new_row, column=col)
                if source.has_style:
                    target.font = source.font.copy()
                    target.border = source.border.copy()
                    target.fill = source.fill.copy()
                    target.number_format = source.number_format
                    target.protection = source.protection.copy()
                    target.alignment = source.alignment.copy()

        # The template has a medium bottom border on C16:F16 that visually
        # separates the parts data from the 'Renunt la garantia' row.
        # After inserting rows the border must sit on the last row before
        # 'Renunt', which is now row 16 + extra_rows.  Move it there.
        _medium = Side(style="medium")
        _none = Side(style=None)
        # Remove bottom from original row 16 (C–F)
        for col_idx in range(3, 7):  # C=3, D=4, E=5, F=6
            cell = sheet.cell(row=16, column=col_idx)
            old = cell.border
            cell.border = Border(left=old.left, top=old.top, right=old.right, bottom=_none)
        # Add medium bottom to the new last data row (16 + extra_rows)
        new_border_row = 16 + extra_rows
        for col_idx in range(3, 7):
            cell = sheet.cell(row=new_border_row, column=col_idx)
            old = cell.border
            cell.border = Border(left=old.left, top=old.top, right=old.right, bottom=_medium)

    # All sections below the rectangle shift by this offset
    row_offset = extra_rows

    # Write ALL defects starting at A14 (no hard limit; section has expanded)
    for i, defect_id in enumerate(defect_ids):
        defect = get_defect_by_id(defect_id)
        if defect:
            sheet[f"A{14 + i}"] = defect["defect_name"]

    # Write ALL client parts starting at C14 (no hard limit)
    for i, part_id in enumerate(part_ids):
        part = get_part_by_id(part_id)
        if part:
            sheet[f"C{14 + i}"] = part["part_name"]

    # A(20+row_offset)..A(24+row_offset+disc_extra): Discovered defects section
    # When n_discovered > 5, extra rows are inserted AFTER the section bottom
    # (at row 25 + row_offset), expanding the box downward.
    # The right-panel text ("Accept lucrarile", "Semnatura client:") sits at
    # rows 20+row_offset and 21+row_offset — ABOVE the insertion point — so it
    # never moves due to the discovered-section expansion.
    discovered_defect_ids = receipt_data.get("discovered_defects", [])
    n_discovered = len(discovered_defect_ids)
    disc_extra_rows = max(0, n_discovered - 5)

    if disc_extra_rows > 0:
        disc_insert_at = 25 + row_offset  # row immediately after section bottom
        sheet.insert_rows(disc_insert_at, amount=disc_extra_rows)

        # Copy formatting from a data row inside the section
        source_disc_row = 20 + row_offset
        for new_row in range(disc_insert_at, disc_insert_at + disc_extra_rows):
            for col in range(1, sheet.max_column + 1):
                source = sheet.cell(row=source_disc_row, column=col)
                target = sheet.cell(row=new_row, column=col)
                if source.has_style:
                    target.font = source.font.copy()
                    target.border = source.border.copy()
                    target.fill = source.fill.copy()
                    target.number_format = source.number_format
                    target.protection = source.protection.copy()
                    target.alignment = source.alignment.copy()

        # Move medium bottom border from old row 24+row_offset to new bottom row
        old_bottom = 24 + row_offset
        new_bottom = old_bottom + disc_extra_rows
        for col_idx in range(1, 7):  # A=1 .. F=6
            cell = sheet.cell(row=old_bottom, column=col_idx)
            old = cell.border
            cell.border = Border(left=old.left, top=old.top, right=old.right, bottom=_none)
        for col_idx in range(1, 7):
            cell = sheet.cell(row=new_bottom, column=col_idx)
            old = cell.border
            cell.border = Border(left=old.left, top=old.top, right=old.right, bottom=_medium)

    # Total offset for all sections below the discovered-defects box
    total_offset = row_offset + disc_extra_rows

    # Write ALL discovered defects (no hard limit; section has expanded)
    for i, defect_id in enumerate(discovered_defect_ids):
        defect = get_defect_by_id(defect_id)
        if defect:
            sheet[f"A{20 + row_offset + i}"] = defect["defect_name"]

    # Labor services starting at B(36 + total_offset)
    labor_ids = receipt_data.get("labor", [])
    total_labor_cost = receipt_data.get("total_labor_cost", 0.0)

    labor_row_base = 36 + total_offset

    for i, labor_id in enumerate(labor_ids):
        labor = get_labor_by_id(labor_id)
        if labor:
            current_row = labor_row_base + i
            if i > 0:
                sheet.insert_rows(current_row)
                for col in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(row=labor_row_base, column=col)
                    target_cell = sheet.cell(row=current_row, column=col)
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()
            sheet[f"B{current_row}"] = labor["service_name"]

    # Add total labor cost at the next row after all labor items
    if labor_ids:
        total_row = labor_row_base + len(labor_ids)
        sheet[f"E{total_row}"] = total_labor_cost
        tva_percentage = get_tva()
        labor_tva = (total_labor_cost * tva_percentage) / (100 + tva_percentage)
        sheet[f"F{total_row}"] = labor_tva

    # Parts Used section: base row 42, shifted by (N_labor-1) inserts + total_offset
    #   N=0 → 42,  N=1 → 42,  N=2 → 43,  N=3 → 44  (then + total_offset)
    billable_parts = receipt_data.get("billable_parts", [])
    total_parts_cost = receipt_data.get("total_parts_cost", 0.0)

    parts_start_row = 42 + max(0, len(labor_ids) - 1) + total_offset

    # Add billable parts starting at the calculated row
    for i, part_data in enumerate(billable_parts):
        current_row = parts_start_row + i
        # If this is not the first part, copy the previous row to maintain formatting
        if i > 0:
            # Copy the parts start row to create a new row with the same formatting
            sheet.insert_rows(current_row)
            for col in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=parts_start_row, column=col)
                target_cell = sheet.cell(row=current_row, column=col)
                # Copy cell formatting
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()

        # Add part data to columns B, C, D, E
        sheet[f"B{current_row}"] = part_data["part_name"]  # Part name
        sheet[f"C{current_row}"] = part_data["units"]  # Units
        sheet[f"D{current_row}"] = part_data["price_per_unit"]  # Price per unit
        # Calculate and add total value (units × price per unit)
        part_total = part_data["units"] * part_data["price_per_unit"]
        sheet[f"E{current_row}"] = part_total
        # Calculate and add TVA for this part at column F (extract TVA from price that includes TVA)
        tva_percentage = get_tva()
        part_tva = (part_total * tva_percentage) / (100 + tva_percentage)
        sheet[f"F{current_row}"] = part_tva

    # Add total parts cost to column E at the next row after all parts
    if billable_parts:
        total_parts_row = parts_start_row + len(billable_parts)
        sheet[f"E{total_parts_row}"] = total_parts_cost
        # Add total TVA for parts at column F (extract TVA from price that includes TVA)
        tva_percentage = get_tva()
        total_parts_tva = (total_parts_cost * tva_percentage) / (100 + tva_percentage)
        sheet[f"F{total_parts_row}"] = total_parts_tva

    # Grand total row (Total Labor + Total Parts) at the next row after total parts.
    # Template base: grand total lives at row 44 (no inserts).
    # Each extra labor row (N-1 inserts for N items) and each extra billable-parts
    # row (M-1 inserts for M items) shifts every row below it down by 1.
    grand_total = total_labor_cost + total_parts_cost
    if billable_parts:
        grand_total_row = total_parts_row + 1
    elif labor_ids:
        # No billable-parts inserts; only labor inserts (N-1) shift base row 44.
        # 44 + (N-1) + total_offset = 43 + N + total_offset
        grand_total_row = 43 + len(labor_ids) + total_offset
    else:
        grand_total_row = 44 + total_offset

    sheet[f"F{grand_total_row}"] = grand_total

    # Executant name row: base 50, shifted by extra labor/parts rows + total_offset
    extra_labor_rows = max(0, len(labor_ids) - 1)
    extra_parts_rows = max(0, len(billable_parts) - 1)
    executant_row = 50 + extra_labor_rows + extra_parts_rows + total_offset
    executant_name = receipt_data.get("executant_name", "")
    if executant_name:
        sheet[f"B{executant_row}"] = executant_name

    # Save the workbook
    workbook.save(output_path)
    workbook.close()

    # Clean up temp logo copy
    if tmp_logo and tmp_logo.exists():
        tmp_logo.unlink(missing_ok=True)

    return str(output_path), warning_messages


def get_template_path() -> Path:
    """Get the path to the template file."""
    return TEMPLATES_DIR / TEMPLATE_FILE


def template_exists() -> bool:
    """Check if the template file exists."""
    return get_template_path().exists()
